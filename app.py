# -*- coding: utf-8 -*-
import os
import time
import unicodedata
import subprocess
from flask import Flask, request, render_template_string, send_from_directory
import json
import openpyxl

# ebcdic package đăng ký thêm các codec EBCDIC SBCS (cp290, cp1027...) mà stdlib
# thieu. Cai qua environment.yml (pip: ebcdic). Import de kich hoat.
try:
    import ebcdic
except ImportError:
    pass

app = Flask(__name__)

# Thư mục lưu trữ các file tải lên và kết quả
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "web_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ==============================================================================
# CÁC HÀM TÁI SỬ DỤNG TỪ ENCODE_TO_EBCDIC.PY
# ==============================================================================
_ENC_MAP = {
    "AS-CP00930": ("cp290", True), "CP00930": ("cp290", True), "CP930": ("cp290", True),
    "AS-CP00939": ("cp1027", True), "CP00939": ("cp1027", True), "CP939": ("cp1027", True),
    "UTF-8": ("utf-8", False), "SHIFT_JIS": ("cp932", False), "WINDOWS-31J": ("cp932", False),
}

def resolve_encoding(name):
    if not name: return "cp290", True, "mac dinh AS-CP00930"
    key = str(name).strip().upper()
    if key in _ENC_MAP: return _ENC_MAP[key][0], _ENC_MAP[key][1], None
    try:
        b"x".decode(name)
        return name, False, "codec python truc tiep (khong coi la EBCDIC)"
    except Exception:
        return None, None, f"encoding khong nhan dien: {name!r}"

def _nib(letter):
    return {"A": 0xA, "B": 0xB, "C": 0xC, "D": 0xD, "E": 0xE, "F": 0xF}.get(str(letter).strip().upper())

def layout_signs(lay):
    ps = lay.get("packed_sign") or {}
    pneg = {_nib(x) for x in (ps.get("negative") or ["B", "D"])} or {0xB, 0xD}
    ppos = {_nib(x) for x in (ps.get("positive") or ["A", "C", "E", "F"])} or {0xC}
    zneg = {_nib(x) for x in (lay.get("zoned_sign_negative") or [])} or {0xD}
    zpos = {_nib(x) for x in (lay.get("zoned_sign_positive") or [])} or {0xC, 0xF}
    return pneg, ppos, zneg, zpos

_REVERSE_DBCS_TABLE = None
def _reverse_dbcs_table():
    """Tải và đảo ngược bảng map CP300 (Unicode -> EBCDIC DBCS)."""
    global _REVERSE_DBCS_TABLE
    if _REVERSE_DBCS_TABLE is None:
        from pathlib import Path
        p = Path(__file__).resolve().parent / "data" / "cp300_dbcs.json"
        try:
            raw = json.loads(p.read_text(encoding="utf-8"))
            _REVERSE_DBCS_TABLE = {v: int(k, 16) for k, v in raw.items()}
        except Exception:
            _REVERSE_DBCS_TABLE = {}
    return _REVERSE_DBCS_TABLE

def encode_char(s: str, length: int, codec: str, is_ebcdic: bool) -> bytes:
    """Mã hóa chuỗi Unicode thành byte EBCDIC, xử lý DBCS và padding."""
    s = s or ""
    dbcs_map = _reverse_dbcs_table()
    padding_char = b'\x40' if is_ebcdic else b'\x20'

    if is_ebcdic and len(s) > 1 and s.startswith(' ') and s.endswith(' '):
        inner_s = s[1:-1]
        out_bytes = bytearray()
        out_bytes.append(0x0E)  # SO

        for char in inner_s:
            if char in dbcs_map:
                dbcs_code = dbcs_map[char]
                out_bytes.append((dbcs_code >> 8) & 0xFF)
                out_bytes.append(dbcs_code & 0xFF)
            else:
                out_bytes.extend(b'\x40\x40')
        
        out_bytes.append(0x0F)  # SI
    else:
        if not is_ebcdic:
            return s.encode(codec, errors='replace').ljust(length, b'\x20')[:length]

        out_bytes = bytearray()
        in_dbcs_mode = False

        for char in s:
            width = unicodedata.east_asian_width(char)
            is_fullwidth = width in ('F', 'W', 'A')

            if is_fullwidth and char in dbcs_map:
                if not in_dbcs_mode:
                    out_bytes.append(0x0E)  # SO
                    in_dbcs_mode = True
                dbcs_code = dbcs_map[char]
                out_bytes.append((dbcs_code >> 8) & 0xFF)
                out_bytes.append(dbcs_code & 0xFF)
            else:
                if in_dbcs_mode:
                    out_bytes.append(0x0F)  # SI
                    in_dbcs_mode = False
                out_bytes.extend(char.encode(codec, errors='replace'))

        if in_dbcs_mode:
            out_bytes.append(0x0F)  # SI

    if len(out_bytes) < length:
        out_bytes.extend(padding_char * (length - len(out_bytes)))

    return bytes(out_bytes[:length])

def encode_packed(s: str, length: int, pneg: set, ppos: set) -> bytes:
    """Mã hóa chuỗi số thành Packed Decimal (COMP-3)."""
    s = (s or "0").strip()
    is_negative = s.startswith('-')
    if is_negative:
        s = s[1:]
    
    digits = "".join(filter(str.isdigit, s))
    
    if len(digits) % 2 == 0:
        digits = '0' + digits

    sign_nibble = next(iter(pneg)) if is_negative else next(iter(ppos))
    
    out_bytes = bytearray()
    for i in range(0, len(digits), 2):
        high = int(digits[i])
        low = int(digits[i+1])
        out_bytes.append((high << 4) | low)

    last_byte = out_bytes[-1]
    out_bytes[-1] = (last_byte & 0xF0) | sign_nibble

    if len(out_bytes) < length:
        return (b'\x00' * (length - len(out_bytes))) + out_bytes
    return bytes(out_bytes[:length])

def encode_zoned(s: str, length: int, zneg: set, zpos: set) -> bytes:
    """Mã hóa chuỗi số thành Zoned Decimal."""
    s = (s or "0").strip()
    is_negative = s.startswith('-')
    if is_negative:
        s = s[1:]

    digits = "".join(filter(str.isdigit, s)).zfill(length)
    
    out_bytes = bytearray()
    for digit in digits:
        out_bytes.append(0xF0 | int(digit))

    if out_bytes:
        sign_nibble = next(iter(zneg)) if is_negative else next(iter(zpos))
        out_bytes[-1] = (sign_nibble << 4) | (out_bytes[-1] & 0x0F)

    return bytes(out_bytes[:length])

HTML_FORM = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Công cụ So sánh Layout</title>
    <style>
        body { font-family: Arial, sans-serif; background-color: #f4f7f6; margin: 40px; }
        .container { max-width: 500px; background: white; margin: auto; padding: 30px; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); }
        h2 { text-align: center; color: #333; margin-bottom: 20px; }
        .form-group { margin-bottom: 15px; }
        label { font-weight: bold; display: block; margin-bottom: 5px; color: #555; }
        input[type="file"], input[type="text"] { width: 100%; padding: 8px; box-sizing: border-box; border: 1px solid #ccc; border-radius: 4px; }
        button { width: 100%; padding: 12px; background-color: #28a745; color: white; border: none; border-radius: 4px; font-size: 16px; cursor: pointer; margin-top: 10px; }
        button:hover { background-color: #218838; }
        .tab { overflow: hidden; border-bottom: 1px solid #ccc; margin-bottom: 20px; }
        .tab button { width: auto; background-color: inherit; float: left; border: none; outline: none; cursor: pointer; padding: 14px 16px; transition: 0.3s; font-size: 17px; border-radius: 4px 4px 0 0; margin: 0; }
        .tab button:hover { background-color: #ddd; }
        .tab button.active { background-color: #ccc; }
        .tabcontent { display: none; }
    </style>
</head>
<body>
    <div class="container">
        <h2>Công Cụ Excel</h2>
        <div class="tab">
            <button class="tablinks" onclick="openTool(event, 'Compare')" id="defaultOpen">So sánh Layout</button>
            <button class="tablinks" onclick="openTool(event, 'Encode')">Encode sang EBCDIC</button>
        </div>

        <div id="Compare" class="tabcontent">
            <h3>So Sánh Layout Data</h3>
            <form action="/compare" method="post" enctype="multipart/form-data">
                <div class="form-group">
                    <label>1. File Excel Testcase (VD: check.xlsx):</label>
                    <input type="file" name="excel_file" accept=".xlsx" required>
                </div>
                <div class="form-group">
                    <label>2. File Data (VD: EBCDIC):</label>
                    <input type="file" name="dat_file" required>
                </div>
                <div class="form-group">
                    <label>3. Bảng mã gốc (VD: cp930, cp290):</label>
                    <input type="text" name="from_enc" value="cp930">
                </div>
                <button type="submit">Bắt đầu so sánh</button>
            </form>
        </div>

        <div id="Encode" class="tabcontent">
            <h3>Encode Excel sang EBCDIC</h3>
            <form action="/encode" method="post" enctype="multipart/form-data">
                <div class="form-group">
                    <label>1. File Excel chứa Layout và Data:</label>
                    <input type="file" name="excel_file" accept=".xlsx" required>
                </div>
                <div class="form-group">
                    <label>2. Bảng mã EBCDIC đích (VD: cp930, cp290):</label>
                    <input type="text" name="to_enc" value="cp930">
                </div>
                <button type="submit">Bắt đầu encode</button>
            </form>
        </div>
    </div>
    <script>
        function openTool(evt, toolName) {
            var i, tabcontent, tablinks;
            tabcontent = document.getElementsByClassName("tabcontent");
            for (i = 0; i < tabcontent.length; i++) {
                tabcontent[i].style.display = "none";
            }
            tablinks = document.getElementsByClassName("tablinks");
            for (i = 0; i < tablinks.length; i++) {
                tablinks[i].className = tablinks[i].className.replace(" active", "");
            }
            document.getElementById(toolName).style.display = "block";
            evt.currentTarget.className += " active";
        }
        document.getElementById("defaultOpen").click();
    </script>
</body>
</html>
"""

HTML_RESULT = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <title>Kết quả so sánh</title>
    <style>
        body { font-family: Arial, sans-serif; background-color: #f4f7f6; margin: 40px; text-align: center; }
        .container { max-width: 95%; background: white; margin: auto; padding: 30px; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); }
        .btn { display: block; width: 80%; margin: 15px auto; padding: 12px; text-decoration: none; color: white; border-radius: 4px; font-weight: bold; }
        .btn-excel { background-color: #1d6f42; } /* Excel green */
        .btn-txt { background-color: #007bff; } /* Text blue */
        .btn-back { background-color: #6c757d; width: auto; display: inline-block; padding: 8px 15px; margin-top: 20px;}
        .log-box { text-align: left; background: #eee; padding: 10px; border-radius: 4px; font-size: 12px; overflow-x: auto; max-height: 200px; }
        .preview-box { text-align: left; background: #272822; color: #f8f8f2; padding: 15px; border-radius: 6px; font-family: Consolas, "Courier New", monospace; font-size: 13px; overflow-x: auto; white-space: pre; max-height: 400px; border: 1px solid #444; }
        .summary-box { display: flex; justify-content: center; gap: 20px; margin: 25px 0; }
        .summary-item { padding: 15px 25px; border-radius: 8px; color: white; font-weight: bold; font-size: 18px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .summary-match { background-color: #28a745; }
        .summary-diff { background-color: #dc3545; }
        .table-container { overflow-x: auto; overflow-y: auto; margin-top: 20px; border: 1px solid #ddd; max-width: 100%; max-height: 600px; }
        .result-table { min-width: 100%; border-collapse: collapse; font-size: 12px; table-layout: fixed; }
        .result-table th, .result-table td { border: 1px solid #ddd; padding: 8px; vertical-align: middle; position: relative; width: 120px; max-width: 120px; height: 40px; box-sizing: border-box; }
        .cell-content { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; width: 100%; display: block; }
        .result-table thead { background-color: #f2f2f2; position: sticky; top: 0; z-index: 2;}
        .result-table th { font-weight: bold; text-align: center; }
        .result-table th .length { font-weight: normal; color: #666; font-size: 11px; }
        .result-table .row-header { font-weight: bold; background-color: #f2f2f2; text-align: center; position: sticky; left: 0; z-index: 3; width: 80px; min-width: 80px; max-width: 80px; }
        .result-table thead .row-header { z-index: 4; }
        .result-table td.status-match { background-color: #e6ffed; }
        .result-table td.status-diff { background-color: #ffebee; }
        .result-table td.has-tooltip { cursor: pointer; }
        .actual-tooltip {
            display: none;
            position: absolute;
            bottom: 100%;
            left: 50%;
            transform: translateX(-50%);
            background-color: #2b2b2b;
            color: #fff;
            padding: 10px 14px;
            border-radius: 6px;
            z-index: 10;
            box-shadow: 0 4px 12px rgba(0,0,0,0.4);
            text-align: left;
            font-family: Consolas, "Courier New", monospace;
            font-size: 13px;
            white-space: nowrap;
        }
        .result-table td.has-tooltip:hover .actual-tooltip {
            display: block;
        }
        .actual-tooltip::after {
            content: "";
            position: absolute;
            top: 100%;
            left: 50%;
            margin-left: -5px;
            border-width: 5px;
            border-style: solid;
            border-color: #2b2b2b transparent transparent transparent;
        }
        .tooltip-row { display: flex; align-items: center; gap: 8px; }
        .tooltip-expected { color: #4ade80; margin-bottom: 8px; padding-bottom: 8px; border-bottom: 1px solid #444; }
        .tooltip-actual { color: #f87171; }
        .tooltip-actual.is-match { color: #4ade80; }
        .tooltip-label { font-size: 12px; color: #aaa; font-family: Arial, sans-serif; width: 115px; flex-shrink: 0; }
        .tooltip-value { background-color: #1a1a1a; padding: 3px 6px; border-radius: 4px; border: 1px solid #444; color: #fff; min-width: 20px; display: inline-block; }
    </style>
</head>
<body>
    <div class="container">
        <h2 style="color: #28a745;">Hoàn tất xử lý!</h2>
        
        {% if summary %}
        <div class="summary-box">
            <div class="summary-item summary-match">✅ Khớp: {{ summary.match }}</div>
            <div class="summary-item summary-diff">❌ Lệch: {{ summary.diff }}</div>
        </div>
        {% endif %}

        {% if detailed_results %}
        <hr style="margin-top:30px; border-top: 1px solid #ccc;">
        <h3 style="text-align: left;">🔎 Chi tiết so sánh (Tổng số: {{ detailed_results.results|length }} dòng):</h3>
        <div class="table-container">
            <table class="result-table">
                <thead>
                    <tr>
                        <th class="row-header">Dòng Excel</th>
                        {% for field in detailed_results.layout %}
                        <th>
                            <div class="cell-content" title="{{ field.name }}">{{ field.name or 'N/A' }}</div>
                            <div class="length">({{ field.len }}b)</div>
                        </th>
                        {% endfor %}
                    </tr>
                </thead>
                <tbody>
                    {% for row in detailed_results.results %}
                    <tr>
                        <td class="row-header">{{ row.row_index }}</td>
                        {% for field in row.fields %}
                        <td class="status-{{ field.status }} {% if field.status in ['diff'] %}has-tooltip{% endif %}">
                            <span class="cell-content">{{ field.expected if field.expected else '""' }}</span>
                            {% if field.status in ['diff'] %}
                                <div class="actual-tooltip">
                                    <div class="tooltip-expected tooltip-row">
                                        <span class="tooltip-label">Dự kiến (Excel):</span>
                                        <span class="tooltip-value">{{ field.expected.replace(' ', '␣') if field.expected else '""' }}</span>
                                    </div>
                                    <div class="tooltip-actual tooltip-row {% if field.status == 'match' %}is-match{% endif %}">
                                        <span class="tooltip-label">Thực tế (Output):</span>
                                        <span class="tooltip-value">{{ field.actual.replace(' ', '␣') if field.actual else '""' }}</span>
                                    </div>
                                </div>
                            {% endif %}
                        </td>
                        {% endfor %}
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% endif %}

        <p>Vui lòng tải các file kết quả bên dưới:</p>
        
        <a href="/download/{{ req_id }}/Result.xlsx" class="btn btn-excel">📥 Tải file so sánh (Result.xlsx)</a>
        
        {% if has_converted %}
        <a href="/download/{{ req_id }}/data_converted.txt" class="btn btn-txt">📥 Tải file Data đã Convert (Shift-JIS)</a>
        {% endif %}
        
        <a href="/" class="btn btn-back">⬅ Quay lại trang chủ</a>

        {% if preview_text %}
        <hr style="margin-top:30px; border-top: 1px solid #ccc;">
        <h3 style="text-align: left;">👀 Preview Dữ liệu Output (5000 byte đầu):</h3>
        <pre class="preview-box">{{ preview_text }}</pre>
        {% endif %}

        <hr style="margin-top:30px; border-top: 1px solid #ccc;">
        <h3 style="text-align: left;">Log xử lý:</h3>
        <pre class="log-box">{{ log }}</pre>
    </div>
</body>
</html>
"""

@app.route('/', methods=['GET'])
def index():
    return render_template_string(HTML_FORM)

@app.route('/compare', methods=['POST'])
def compare():
    excel_file = request.files.get('excel_file')
    dat_file = request.files.get('dat_file')
    from_enc = request.form.get('from_enc', 'cp930').strip()

    if not all([excel_file, dat_file]):
        return "Vui lòng chọn đủ file Excel và file Data!", 400

    # Tạo một thư mục riêng biệt cho mỗi lần chạy (dựa trên Timestamp)
    req_id = str(int(time.time() * 1000))
    req_dir = os.path.join(UPLOAD_DIR, req_id)
    os.makedirs(req_dir, exist_ok=True)

    excel_path = os.path.join(req_dir, "check.xlsx")
    dat_path = os.path.join(req_dir, "data.dat")
    result_excel_path = os.path.join(req_dir, "Result.xlsx")
    result_json_path = os.path.join(req_dir, "result_detail.json")
    
    excel_file.save(excel_path)
    dat_file.save(dat_path)

    # Gọi script python hiện tại của bạn
    cmd = ["python", "compare_fixed_length.py", "--excel", excel_path, "--dat", dat_path, "--from_enc", from_enc, "--out_excel", result_excel_path, "--out_json", result_json_path]

    env = os.environ.copy()
    env["PYTHONUTF8"] = "1"
    process = subprocess.run(cmd, capture_output=True, text=True, cwd=BASE_DIR, encoding='utf-8', errors='replace', env=env)

    has_converted = os.path.exists(os.path.join(req_dir, "data_converted.txt"))
    log_output = process.stdout + "\n" + process.stderr

    # Đọc nội dung file để preview (ưu tiên file đã convert)
    # Vì giờ không còn file convert, ta sẽ đọc file .dat gốc và hiển thị dạng hex
    preview_text = ""
    try:
        # Đọc file .dat gốc và hiển thị dạng hex
        with open(dat_path, 'rb') as f:
            preview_bytes = f.read(5000)
            preview_text = preview_bytes.hex(' ', 16)
    except Exception as e:
        preview_text = f"Không thể tải trước preview: {e}"

    # Parse summary từ log
    summary = None
    for line in log_output.splitlines():
        if line.startswith("[SUMMARY]"):
            try:
                parts = line.replace("[SUMMARY]", "").strip().split(',')
                match_count = int(parts[0].split(':')[1].strip())
                diff_count = int(parts[1].split(':')[1].strip())
                summary = {"match": match_count, "diff": diff_count}
                break # Tìm thấy là dừng
            except (IndexError, ValueError):
                pass # Bỏ qua nếu dòng summary bị lỗi
    
    # Đọc detailed results từ file JSON (ưu tiên), nếu không có mới parse từ log
    detailed_results = None
    if os.path.exists(result_json_path):
        try:
            with open(result_json_path, "r", encoding="utf-8") as f_json:
                detailed_results = json.load(f_json)
        except Exception as e:
            print(f"Lỗi khi đọc file json chi tiết: {e}")

    if not detailed_results:
        for line in log_output.splitlines():
            if line.startswith("[RESULT_JSON]"):
                try:
                    json_str = line.replace("[RESULT_JSON]", "", 1)
                    detailed_results = json.loads(json_str)
                except json.JSONDecodeError:
                    pass # Bỏ qua nếu JSON bị lỗi
                break

    return render_template_string(
        HTML_RESULT, 
        req_id=req_id, 
        has_converted=has_converted,
        log=log_output,
        preview_text=preview_text,
        summary=summary,
        detailed_results=detailed_results
    )


@app.route('/download/<req_id>/<filename>')
def download(req_id, filename):
    # Tìm file convert động
    req_dir = os.path.join(UPLOAD_DIR, req_id)
    converted_file = None
    if os.path.exists(req_dir):
        try:
            # Tìm file convert của chức năng "So sánh Layout"
            converted_file = next((f for f in os.listdir(req_dir) if f.endswith('_converted.txt')), None)
        except (StopIteration, FileNotFoundError):
            pass # Bỏ qua nếu không tìm thấy, vì có thể đây là request của chức năng "Encode"

    # Thêm 'output.dat' vào danh sách các file được phép tải
    allowed_files = ["Result.xlsx", "output.dat"]
    if converted_file: allowed_files.append(converted_file) # Cho phép tải file convert
    if filename not in allowed_files:
        return "File không hợp lệ!", 400
        
    directory = os.path.join(UPLOAD_DIR, req_id)
    if not os.path.exists(os.path.join(directory, filename)):
        return "Không tìm thấy file kết quả. Có thể đã xảy ra lỗi trong lúc chạy.", 404
        
    return send_from_directory(directory, filename, as_attachment=True)

@app.route('/encode', methods=['POST'])
def encode():
    excel_file = request.files.get('excel_file')
    to_enc = request.form.get('to_enc', 'cp930').strip()

    if not excel_file:
        return "Vui lòng chọn file Excel!", 400

    req_id = str(int(time.time() * 1000))
    req_dir = os.path.join(UPLOAD_DIR, req_id)
    os.makedirs(req_dir, exist_ok=True)

    excel_path = os.path.join(req_dir, "source.xlsx")
    dat_path = os.path.join(req_dir, "output.dat")
    excel_file.save(excel_path)

    log_output = ""
    try:
        log_output += "1. Đang đọc file Excel...\n"
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        log_output += "2. Đang đọc layout từ Excel...\n"
        fields = []
        if sheet.max_row < 3:
            raise ValueError("Lỗi: File Excel phải có ít nhất 3 dòng (dòng 1: Name, dòng 2: Type, dòng 3: Length).")

        field_names = [cell.value for cell in sheet[1]]
        field_types = [cell.value for cell in sheet[2]]
        field_lengths = [cell.value for cell in sheet[3]]

        for i in range(sheet.max_column):
            length = field_lengths[i]
            if length is not None and str(length).strip().isdigit():
                fields.append({
                    'name': str(field_names[i] or f"F_{i+1}").strip(), 
                    'type': str(field_types[i] or 'char').strip().lower(),
                    'len': int(length), 
                    'col': i + 1
                })

        if not fields:
            raise ValueError("Lỗi: Không đọc được layout từ Excel.")

        codec, is_ebcdic, enc_note = resolve_encoding(to_enc)
        if enc_note: log_output += f" -> [Encoding Note] {enc_note}\n"
        if not codec:
            raise ValueError(f"Lỗi: {enc_note}")

        pneg, ppos, zneg, zpos = layout_signs({})
        log_output += f" -> Layout: {len(fields)} fields, encoding={codec}\n"

        all_records_bytes = bytearray()
        start_data_row = 4

        log_output += "3. Đang xử lý các dòng dữ liệu...\n"
        for row_idx in range(start_data_row, sheet.max_row + 1):
            record_bytes = bytearray()
            has_data_in_row = any(sheet.cell(row=row_idx, column=field['col']).value is not None for field in fields)
            
            if has_data_in_row:
                for field in fields:
                    cell_val = sheet.cell(row=row_idx, column=field['col']).value
                    val_str = str(cell_val) if cell_val is not None else ""
                    
                    if field['type'] == 'packed':
                        encoded_chunk = encode_packed(val_str, field['len'], pneg, ppos)
                    elif field['type'] == 'zoned':
                        encoded_chunk = encode_zoned(val_str, field['len'], zneg, zpos)
                    else: # Mặc định là 'char'
                        encoded_chunk = encode_char(val_str, field['len'], codec, is_ebcdic)
                    record_bytes.extend(encoded_chunk)
                all_records_bytes.extend(record_bytes)
                log_output += f" - Đã xử lý dòng {row_idx}, tạo ra {len(record_bytes)} bytes.\n"

        log_output += f"4. Đang ghi kết quả ra file: {dat_path}\n"
        with open(dat_path, 'wb') as f_out:
            f_out.write(all_records_bytes)
        log_output += "--- Hoàn tất ---\n"

    except Exception as e:
        log_output += f"\nLỖI XẢY RA: {e}\n"

    # Tạo trang kết quả đơn giản cho chức năng encode
    return f"""
    <!DOCTYPE html>
    <html lang="vi">
    <head><title>Kết quả Encode</title>
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f4f7f6; margin: 40px; }}
            .container {{ max-width: 800px; background: white; margin: auto; padding: 30px; border-radius: 8px; box-shadow: 0 4px 8px rgba(0,0,0,0.1); }}
            .btn {{ display: inline-block; padding: 12px 20px; text-decoration: none; color: white; border-radius: 4px; font-weight: bold; }}
            .btn-dat {{ background-color: #007bff; }}
            .btn-back {{ background-color: #6c757d; }}
            .log-box {{ text-align: left; background: #eee; padding: 10px; border-radius: 4px; font-size: 12px; overflow-x: auto; white-space: pre-wrap; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h2>Kết quả Encode sang EBCDIC</h2>
            <a href="/download/{req_id}/output.dat" class="btn btn-dat">📥 Tải file EBCDIC (.dat)</a>
            <a href="/" class="btn btn-back">⬅ Quay lại trang chủ</a>
            <hr style="margin: 20px 0;">
            <h3>Log xử lý:</h3>
            <pre class="log-box">{log_output}</pre>
        </div>
    </body>
    </html>
    """

if __name__ == '__main__':
    # Mở mạng LAN (0.0.0.0) để người khác cùng IP truy cập được
    print("Đang khởi động Web Server... Truy cập vào http://127.0.0.1:5000 trên trình duyệt.")
    app.run(host='0.0.0.0', port=5000, debug=True)