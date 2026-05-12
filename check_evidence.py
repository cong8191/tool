# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill
import os
import argparse
import re

def format_val(val):
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        res = str(int(val))
    else:
        res = str(val).strip()
        
    # Chuyển đổi Full-width sang Half-width để đồng bộ dữ liệu
    half_res = ""
    for c in res:
        code = ord(c)
        if 0xFF01 <= code <= 0xFF5E:
            half_res += chr(code - 0xFEE0)
        elif code == 0x3000: # Khoảng trắng Full-width
            half_res += ' '
        else:
            half_res += c
    return half_res

def check_evidence(file_path):
    if not os.path.exists(file_path):
        print(f"Lỗi: Không tìm thấy file:\n{file_path}")
        return

    print(f"Đang đọc file: {file_path}")
    try:
        # Bỏ data_only=True để có thể lưu lại (save) file cùng với format màu sắc
        wb = openpyxl.load_workbook(file_path)
        
        sheet_plan_name = 'テスト計画書兼結果報告書(マッピング)'
        sheet_evi_name = 'エビデンス(マッピング)'
        
        if sheet_plan_name not in wb.sheetnames:
            print(f"Lỗi: Không tìm thấy sheet '{sheet_plan_name}'")
            return
            
        if sheet_evi_name not in wb.sheetnames:
            print(f"Lỗi: Không tìm thấy sheet '{sheet_evi_name}'")
            return

        ws_plan = wb[sheet_plan_name]
        ws_evi = wb[sheet_evi_name]

        print(f"1. Đang trích xuất danh sách Testcase từ sheet '{sheet_plan_name}'...")
        testcases = []
        tc_dict = {} # Lưu trữ thông tin field test/không test: { '1-1-1': {'gray_in': set(), 'test_in': set(), ...} }
        
        last_b = ""
        last_c = ""
        last_d = ""
        
        all_in_fields = set()
        all_out_fields = set()
        for col in range(8, ws_plan.max_column + 1):
            in_name = format_val(ws_plan.cell(row=3, column=col).value).replace('\n', '').strip()
            out_name = format_val(ws_plan.cell(row=5, column=col).value).replace('\n', '').strip()
            if in_name: all_in_fields.add(in_name)
            if out_name: all_out_fields.add(out_name)
        
        # Theo cấu trúc chuẩn, testcase thường bắt đầu từ dòng 8 hoặc 10 trở đi
        # Cột H (index 8) trở đi chứa các field (input/output)
        for row in range(8, ws_plan.max_row + 1):
            val_b = format_val(ws_plan.cell(row=row, column=2).value)
            val_c = format_val(ws_plan.cell(row=row, column=3).value)
            val_d = format_val(ws_plan.cell(row=row, column=4).value)
            
            # Xử lý Merge cell: Giữ lại giá trị của các ô bị gộp, tự động làm sạch nếu qua cụm mới
            if val_b:
                last_b = val_b
                if not val_c: last_c = ""
                if not val_d: last_d = ""
            if val_c:
                last_c = val_c
                if not val_d: last_d = ""
            if val_d:
                last_d = val_d
                
            # Ghép B-C-D thành ID testcase (VD: 1-1-1)
            if last_b and last_c and last_d:
                tc_id = f"{last_b}-{last_c}-{last_d}"
                
                # Bỏ qua các dòng Header (nếu có chứa chữ 項目, No...)
                if "項目" not in tc_id and "No" not in tc_id:
                    has_data = False
                    test_in = set()
                    test_out = set()
                    
                    for col in range(8, ws_plan.max_column + 1):
                        cell_val = format_val(ws_plan.cell(row=row, column=col).value)
                        
                        # Lấy chính xác tên field ở dòng 3 (Input) và dòng 5 (Output)
                        in_name = format_val(ws_plan.cell(row=3, column=col).value).replace('\n', '').strip()
                        out_name = format_val(ws_plan.cell(row=5, column=col).value).replace('\n', '').strip()
                        
                        if not in_name and not out_name:
                            continue
                            
                        # Coi tất cả các ô có data (mock data, ○, text...) là ĐƯỢC TEST
                        if cell_val and cell_val not in ["-", "対象外", "検証不可", ""]:
                            has_data = True
                            if in_name: test_in.add(in_name)
                            if out_name: test_out.add(out_name)
                    
                    if has_data:
                        if tc_id not in testcases:
                            testcases.append(tc_id)
                            tc_dict[tc_id] = {'test_in': test_in, 'test_out': test_out}
                        else:
                            tc_dict[tc_id]['test_in'].update(test_in)
                            tc_dict[tc_id]['test_out'].update(test_out)
                        
        print(f" -> Tìm thấy {len(testcases)} testcases cần kiểm tra.")
        
        # print("\n--- DANH SÁCH FIELD ĐƯỢC TEST (○) TỪNG TESTCASE ---")
        # for tc in testcases:
        #     t_in = tc_dict[tc]['test_in']
        #     t_out = tc_dict[tc]['test_out']
        #     print(f"[{tc}]")
        #     if t_in:
        #         print(f"  + Input : {', '.join(t_in)}")
        #     if t_out:
        #         print(f"  + Output: {', '.join(t_out)}")
        # print("---------------------------------------------------\n")

        print(f"2. Đang đối chiếu và tô màu xám các field không test trong sheet '{sheet_evi_name}'...")
        missing_tcs = []
        found_tcs = []
        
        cells_colored = 0
        
        evi_in_tcs = set()
        evi_out_tcs = set()
        
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # --- BƯỚC 1: BÔI XÁM CÁC FIELD KHÔNG ĐƯỢC TEST ---
        for row in range(1, ws_evi.max_row + 1):
            for c in range(1, ws_evi.max_column + 1):
                val_c = format_val(ws_evi.cell(row=row, column=c).value)
                
                # Cứ thấy chữ "並び順" là xác định đây là một bảng Data
                if '並び順' in val_c:
                    # --- Xác định bảng này thuộc Input hay Output ---
                    table_headers = []
                    for hc in range(c + 1, ws_evi.max_column + 1):
                        val_hc_check = format_val(ws_evi.cell(row=row, column=hc).value)
                        if '並び順' in val_hc_check: break
                        for r_idx in range(row + 1, row + 6):
                            h_val = format_val(ws_evi.cell(row=r_idx, column=hc).value).replace('\n', '').strip()
                            if h_val: table_headers.append(h_val)
                            
                    in_count = 0
                    out_count = 0
                    for h_val in table_headers:
                        is_in = any(h_val == tf or tf in h_val or h_val in tf for tf in all_in_fields)
                        is_out = any(h_val == tf or tf in h_val or h_val in tf for tf in all_out_fields)
                        if is_in and not is_out: in_count += 1
                        if is_out and not is_in: out_count += 1
                        
                    table_is_input = (in_count >= out_count)
                    table_is_output = (out_count >= in_count)

                    # Quét các cột bên phải chữ 並び順
                    for hc in range(c + 1, ws_evi.max_column + 1):
                        val_hc_check = format_val(ws_evi.cell(row=row, column=hc).value)
                        # Nếu đụng trúng một bảng "並び順" khác cùng dòng -> Dừng quét cột để sang bảng kia
                        if '並び順' in val_hc_check:
                            break 
                            
                        # Quét toàn bộ các dòng header (từ row+1 đến row+5) để tìm Tên Field
                        header_vals = []
                        for r_idx in range(row + 1, row + 6):
                            if r_idx > ws_evi.max_row: break
                            h_val = format_val(ws_evi.cell(row=r_idx, column=hc).value).replace('\n', '').strip()
                            if h_val:
                                header_vals.append(h_val)
                        
                        # Nếu không có bất kỳ Header nào, check xem vùng data bên dưới có border không
                        if not header_vals:
                            check_data_cell = ws_evi.cell(row=row + 6, column=hc)
                            has_top_border = False
                            if check_data_cell.border:
                                has_top_border = bool(
                                    (check_data_cell.border.left and check_data_cell.border.left.style) or
                                    (check_data_cell.border.right and check_data_cell.border.right.style) or
                                    (check_data_cell.border.top and check_data_cell.border.top.style) or
                                    (check_data_cell.border.bottom and check_data_cell.border.bottom.style)
                                )
                            if not has_top_border:
                                continue
                            
                        offset = 6
                        last_row_tcs = []
                        while True:
                            curr_row = row + offset
                            if curr_row > ws_evi.max_row: break
                            
                            data_cell = ws_evi.cell(row=curr_row, column=hc)
                            
                            # Check xem đã hết bảng data chưa (cột 並び順 rỗng VÀ ô data không có viền)
                            if offset > 6:
                                sort_val = format_val(ws_evi.cell(row=curr_row, column=c).value)
                                has_border = False
                                if data_cell.border:
                                    has_border = bool(
                                        (data_cell.border.left and data_cell.border.left.style) or
                                        (data_cell.border.right and data_cell.border.right.style) or
                                        (data_cell.border.top and data_cell.border.top.style) or
                                        (data_cell.border.bottom and data_cell.border.bottom.style)
                                    )
                                if not sort_val and not has_border: 
                                    break
                            
                            # Tìm Testcase ID trên dòng data (cột B hoặc C)
                            tc_val = format_val(ws_evi.cell(row=curr_row, column=2).value)
                            if not tc_val:
                                tc_val = format_val(ws_evi.cell(row=curr_row, column=3).value)
                                
                            row_tcs = re.findall(r'\d+-\d+-\d+', tc_val)
                            if not row_tcs:
                                row_tcs = last_row_tcs # Kế thừa từ dòng trên nếu ô bị merge
                            else:
                                last_row_tcs = row_tcs
                                
                            # Lưu Testcase ID vừa quét được vào danh sách Input/Output của Evidence
                            if table_is_input:
                                evi_in_tcs.update(row_tcs)
                            if table_is_output:
                                evi_out_tcs.update(row_tcs)
                                
                            tested_fields = set()
                            for t_id in row_tcs:
                                if t_id in tc_dict:
                                    tested_fields.update(tc_dict[t_id]['test_in'])
                                    tested_fields.update(tc_dict[t_id]['test_out'])
                                    
                            is_tested = False
                            for h_val in header_vals:
                                if h_val in tested_fields:
                                    is_tested = True
                                    break
                                for tf in tested_fields:
                                    if tf in h_val or h_val in tf:
                                        is_tested = True
                                        break
                                if is_tested:
                                    break
                            
                            if is_tested:
                                data_cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                            else:
                                data_cell.fill = gray_fill
                                cells_colored += 1
                            offset += 1

        # --- Cảnh báo Testcase thiếu mapping hai chiều trong sheet Evidence ---
        evi_warnings = []
        all_evi_tcs = evi_in_tcs.union(evi_out_tcs)
        for tc in sorted(list(all_evi_tcs)):
            if tc not in tc_dict: continue # Bỏ qua các ID lạ không có trong Test Plan
            
            if tc in evi_in_tcs and tc not in evi_out_tcs:
                evi_warnings.append(f"Testcase [{tc}] có trong Evidence Input nhưng thiếu Evidence Output.")
            elif tc in evi_out_tcs and tc not in evi_in_tcs:
                evi_warnings.append(f"Testcase [{tc}] có trong Evidence Output nhưng thiếu Evidence Input.")

        if evi_warnings:
            print("\n[CẢNH BÁO] Phát hiện Testcase thiếu mapping hai chiều (trong sheet Evidence):")
            for w in evi_warnings:
                print(f"  ! {w}")

        print(f"3. Đang lưu lại các thay đổi...")
        wb.save(file_path)
        
        print("\n================ KẾT QUẢ KIỂM TRA ================")
        print(f"🎨 Đã tự động bôi xám: {cells_colored} ô data không được kiểm tra.")
        print("==================================================")
            
    except Exception as e:
        print(f"Lỗi khi xử lý file: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Kiểm tra xem Testcase đã có mặt trong Evidence hay chưa")
    parser.add_argument("file", nargs="?", default=r"D:\Project\151_ISA_AsteriaWrap\trunk\04_Testcase\E302\SHRF0111\単体テスト仕様書兼成績書_E302_SHRF0111.xlsx", help="Đường dẫn file excel")
    args = parser.parse_args()
    
    check_evidence(args.file)