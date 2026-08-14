# -*- coding: utf-8 -*-
import argparse
import openpyxl
from openpyxl.styles import PatternFill
import os
import sys
from pathlib import Path
import json

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# ==============================================================================
# CÁC HÀM TÁI SỬ DỤNG TỪ DECODE_TOOL.PY
# ==============================================================================
try:
    import ebcdic
except ImportError:
    pass # Sẽ xử lý ở dưới

# Bảng mã EBCDIC cơ sở (SBCS - Single-Byte)
_ENC_MAP = {
    "AS-CP00930": ("cp290", True), "CP00930": ("cp290", True), "CP930": ("cp290", True),
    "AS-CP00939": ("cp1027", True), "CP00939": ("cp1027", True), "CP939": ("cp1027", True),
    "UTF-8": ("utf-8", False), "SHIFT_JIS": ("cp932", False), "WINDOWS-31J": ("cp932", False),
}

def resolve_encoding(name):
    if not name: return "cp290", True, "mac dinh AS-CP00930"
    key = str(name).strip().upper()
    if key in _ENC_MAP: return _ENC_MAP[key][0], _ENC_MAP[key][1], None
    try:
        b"x".decode(name)
        return name, False, "codec python truc tiep (khong coi la EBCDIC)"
    except Exception:
        return None, None, f"encoding khong nhan dien: {name!r}"

def _nib(letter):
    return {"A": 0xA, "B": 0xB, "C": 0xC, "D": 0xD, "E": 0xE, "F": 0xF}.get(str(letter).strip().upper())

def layout_signs(lay):
    ps = lay.get("packed_sign") or {}
    pneg = {_nib(x) for x in (ps.get("negative") or ["B", "D"])} or {0xB, 0xD}
    zneg = {_nib(x) for x in (lay.get("zoned_sign_negative") or [])} or {0xD}
    return pneg, zneg

def _fmt_number(digit_str, neg, digits, decimals):
    s = digit_str
    if digits and digits > 0: s = s[-digits:].zfill(digits)
    if decimals and decimals > 0:
        ip = s[:-decimals] or "0"
        fp = s[-decimals:]
        s = f"{ip}.{fp}"
    return f"-{s}" if neg else s

def decode_zoned(b, digits, decimals, neg_nibbles):
    if not b: return ""
    digs = "".join(str(x & 0x0F) for x in b)
    sign = (b[-1] >> 4) & 0x0F
    neg = sign in neg_nibbles
    return _fmt_number(digs, neg, digits, decimals)

def decode_packed(b, digits, decimals, neg_nibbles):
    if not b: return ""
    nib = []
    for x in b:
        nib.extend([(x >> 4) & 0x0F, x & 0x0F])
    sign = nib.pop()
    digs = "".join(map(str, nib))
    neg = sign in neg_nibbles
    return _fmt_number(digs, neg, digits, decimals)

_DBCS_TABLE = None
def _dbcs_table():
    """Tải bảng map CP300 (DBCS Kanji) từ file JSON."""
    global _DBCS_TABLE
    if _DBCS_TABLE is None:
        p = Path(__file__).resolve().parent / "data" / "cp300_dbcs.json"
        try:
            raw = json.loads(p.read_text(encoding="utf-8"))
            _DBCS_TABLE = {int(k, 16): v for k, v in raw.items()}
        except Exception:
            _DBCS_TABLE = {}
    return _DBCS_TABLE

def decode_char(b, codec, is_ebcdic, warns):
    """Giải mã ký tự EBCDIC, xử lý DBCS (Kanji) bằng bảng map."""
    if not is_ebcdic or 0x0E not in b:
        try:
            return b.decode(codec, errors="replace")
        except LookupError:
            warns.add(f"codec '{codec}' không có (cai 'ebcdic'?)")
            return f"<?codec:{b.hex()}>"
    tbl = _dbcs_table()
    if not tbl:
        warns.add("Không tải được bảng CP300 (data/cp300_dbcs.json) -> Kanji sẽ ra placeholder.")
    out = []
    i, n = 0, len(b)
    while i < n:
        c = b[i]
        if b[i] == 0x0E:
            out.append(" ") # Thay thế 0x0E bằng space
            i += 1
            while i < n and b[i] != 0x0F:
                if i + 1 < n and b[i+1] != 0x0F:
                    key = (b[i] << 8) | b[i+1]
                    ch = tbl.get(key, f"《{key:04X}》")
                    out.append(ch)
                    i += 2
                else: # Lẻ byte
                    out.append(f"《?{b[i]:02X}》"); i += 1
            if i < n and b[i] == 0x0F:
                out.append(" ") # Thay thế 0x0F bằng space
                i += 1
        elif c == 0x0F: # SI lẻ
            out.append(" ") # Thay thế 0x0F lẻ bằng space
            i += 1
        else:
            try:
                out.append(bytes([c]).decode(codec, errors="replace"))
            except:
                out.append("?")
            i += 1
    return "".join(out)

def split_records(data, reclen):
    if not reclen or not data: return []
    return [data[i:i + reclen] for i in range(0, len(data), reclen)]

# ==============================================================================

def main():
    parser = argparse.ArgumentParser(description="Tool tìm kiếm data trong file output dựa trên tổng số byte và đánh dấu trực tiếp vào file Excel.")
    parser.add_argument("--excel", required=True, help="Đường dẫn file Excel (Dòng 1: Name, Dòng 2: Length, Dòng 3 trở đi: Data testcase).")
    parser.add_argument("--dat", required=True, help="Đường dẫn file data EBCDIC (.dat).")
    # Bỏ layout JSON, thêm lại from_enc và to_enc
    parser.add_argument("--from_enc", default="cp930", help="Bảng mã EBCDIC nguồn (VD: cp930, cp290). Mặc định là cp930.")
    parser.add_argument("--to_enc", default="cp932", help="Bảng mã để hiển thị (VD: cp932, utf-8). Mặc định là cp932.")
    parser.add_argument("--out_excel", default="Result.xlsx", help="Đường dẫn file Excel xuất ra báo cáo (Mặc định: Result.xlsx).")
    parser.add_argument("--out_json", default=None, help="Đường dẫn file JSON chi tiết.")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"Lỗi: Không tìm thấy file excel '{args.excel}'")
        return
        
    if not os.path.exists(args.dat):
        print(f"Lỗi: Không tìm thấy file data '{args.dat}'")
        return
        
    print("1. Đang đọc file Excel...")
    wb = openpyxl.load_workbook(args.excel)
    sheet = wb.active

    print("2. Đang đọc layout từ Excel (dòng 1: Name, dòng 2: Length)...")
    fields = []
    reclen = 0
    if sheet.max_row < 2:
        print("Lỗi: File Excel phải có ít nhất 2 dòng (dòng 1 cho Name, dòng 2 cho Length).")
        return

    field_names = [cell.value for cell in sheet[1]]
    field_lengths = [cell.value for cell in sheet[2]]

    for i in range(sheet.max_column):
        name = field_names[i]
        length = field_lengths[i]
        
        if length is not None and str(length).strip().isdigit():
            length = int(length)
            fields.append({
                'name': str(name).strip() if name else f"FIELD_{i+1}", 
                'len': length, 
                'col': i + 1
            })
            reclen += length

    if not fields or not reclen:
        print("Lỗi: Không đọc được layout từ Excel hoặc tổng chiều dài bằng 0.")
        return

    # Sử dụng from_enc để xác định encoding
    codec, is_ebcdic, enc_note = resolve_encoding(args.from_enc)
    if enc_note: print(f" -> [Encoding Note] {enc_note}")
    if not codec:
        print(f"Lỗi: {enc_note}")
        return

    # Sử dụng cấu hình mặc định cho packed/zoned signs
    pneg, zneg = {0xB, 0xD}, {0xD}
    warns = set()

    print(f" -> Layout: {len(fields)} fields, record_length={reclen}, encoding={codec}")

    print("3. Đang đọc file data EBCDIC...")
    dat_bytes = open(args.dat, 'rb').read()
    records = split_records(dat_bytes, reclen)
    print(f" -> File data có {len(records)} record.")

    print("4. Đang so sánh và đánh dấu vào Excel...")
    fill_match = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Xanh lá (Khớp)
    fill_diff = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Đỏ (Lệch)

    # Chuẩn bị file convert
    dat_basename = os.path.splitext(os.path.basename(args.dat))[0]
    converted_txt_path = os.path.join(os.path.dirname(args.out_excel), f"{dat_basename}_converted.txt")
    converted_lines = []

    # Tìm dòng bắt đầu của data trong Excel (dòng đầu tiên sau 2 dòng header)
    start_data_row = 3 
    if sheet.max_row < start_data_row:
        print("Lỗi: Không có dòng data nào để kiểm tra (cần data từ dòng 3 trở đi).")
        return

    results_data = []
    match_count = 0
    diff_count = 0

    # Tính tổng số record cần xử lý (không giới hạn theo sheet.max_row)
    excel_data_rows = max(0, sheet.max_row - start_data_row + 1)
    total_records = max(len(records), excel_data_rows)
    print(f" -> Tổng số dòng/record xử lý: {total_records} (DAT: {len(records)} record, Excel: {excel_data_rows} dòng data)")

    for i in range(total_records):
        excel_row_idx = start_data_row + i
        rec_bytes = records[i] if i < len(records) else b""

        if i >= len(records):
            warns.add(f"File Excel có nhiều dòng hơn số record trong file DAT (Dòng {excel_row_idx} không có dữ liệu DAT tương ứng).")
        elif excel_row_idx > sheet.max_row:
            warns.add(f"File DAT có nhiều record hơn số dòng trong file Excel (Record {i+1} không có dòng Excel tương ứng).")

        print(f" - Đang kiểm tra Record {i+1} (Excel dòng {excel_row_idx})")

        row_result = {
            'row_index': excel_row_idx,
            'fields': []
        }

        record_decoded_parts = []
        # Từ khóa để nhận diện trường số
        NUMERIC_KEYWORDS = ["NUM", "DEC", "PACKED", "ZONED", "AMT", "KINGAKU", "SURYO", "NUMBER", "COUNT", "KOSUU", "TANI", "GAKU"]

        offset = 0
        for field in fields:
            field_len = field.get("len", 0)
            field_name = field.get("name", "").upper()
            chunk = rec_bytes[offset:offset + field_len] if offset < len(rec_bytes) else b""
            offset += field_len

            # Tự động nhận diện kiểu dữ liệu
            is_potentially_numeric = any(keyword in field_name for keyword in NUMERIC_KEYWORDS)

            decoded_values = {}
            # Luôn giải mã kiểu char
            decoded_values['char'] = decode_char(chunk, codec, is_ebcdic, warns)

            # Nếu có khả năng là số, thử giải mã thêm packed và zoned
            if is_potentially_numeric:
                # Giả định mặc định cho digits và decimals khi không có layout
                decoded_values['packed'] = decode_packed(chunk, None, 0, pneg)
                decoded_values['zoned'] = decode_zoned(chunk, None, 0, zneg)
            
            # Giá trị mặc định để so sánh
            actual_str = decoded_values['char']
            record_decoded_parts.append(actual_str)

            # Lấy dữ liệu mong đợi từ file Excel
            cell = sheet.cell(row=excel_row_idx, column=field['col'])
            expected_val = cell.value
            expected_str = ""
            has_data = False

            if expected_val is not None:
                has_data = True
                expected_str = str(expected_val)
            
            # So sánh và tô màu
            status = 'empty' # Mặc định là ô trống trong Excel
            if has_data:
                # So sánh với tất cả các khả năng đã giải mã
                is_match = False
                for dec_type, dec_val in decoded_values.items():
                    if expected_str.strip() == dec_val.strip():
                        is_match = True
                        actual_str = dec_val # Cập nhật actual_str thành giá trị khớp
                        break
                if is_match:
                    cell.fill = fill_match
                    match_count += 1
                    status = 'match'
                else:
                    cell.fill = fill_diff
                    diff_count += 1
                    status = 'diff'
            else:
                # Excel không có dữ liệu (ô trống)
                if i < len(records) and actual_str.strip() != "":
                    # DAT có dữ liệu mà Excel trống -> Diff
                    cell.fill = fill_diff
                    diff_count += 1
                    status = 'diff'
                else:
                    # Cả 2 đều trống
                    status = 'empty'
            
            row_result['fields'].append({
                'name': field.get('name'),
                'expected': expected_str,
                'actual': actual_str,
                'status': status
            })
        results_data.append(row_result)
        converted_lines.append("".join(record_decoded_parts))

    # In ra JSON để web UI có thể đọc và hiển thị chi tiết
    json_payload = {'layout': fields, 'results': results_data}
    if args.out_json:
        try:
            with open(args.out_json, "w", encoding="utf-8") as f_json:
                json.dump(json_payload, f_json, ensure_ascii=False)
            print(f" -> Đã lưu JSON chi tiết thành công tại: {args.out_json}")
        except Exception as e:
            print(f" -> Lỗi khi lưu JSON chi tiết: {e}")

    print(f"[RESULT_JSON]{json.dumps(json_payload, ensure_ascii=False)}")
    print(f"[SUMMARY] Match: {match_count}, Diff: {diff_count}")

    # Ghi file đã convert
    try:
        with open(converted_txt_path, "w", encoding="cp932", errors="replace") as f:
            f.write("\n".join(converted_lines))
        print(f" -> Đã tạo file convert thành công tại: {converted_txt_path}")
    except Exception as e:
        print(f" -> Lỗi khi tạo file convert: {e}")

    if warns:
        print("\nCảnh báo trong quá trình xử lý:")
        for w in sorted(list(warns)): print(f" - {w}")

    print(f"\n5. Đang lưu kết quả ra: {args.out_excel}")
    wb.save(args.out_excel)
    print("--- Hoàn tất ---")

if __name__ == "__main__":
    main()