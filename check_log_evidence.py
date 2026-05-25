# -*- coding: utf-8 -*-
import openpyxl
import argparse
import os
import datetime
import re
import json
import shutil
import sys

# Imports for image processing, needed by both manual Gemini and Tesseract
try:
    from PIL import Image, ImageDraw
    import io
    HAS_IMAGE_LIBS = True
except ImportError:
    HAS_IMAGE_LIBS = False

# Flag for manual mode
MANUAL_GEMINI_WEB = True # Đổi thành True để dùng Gemini web miễn phí, False để dùng Tesseract OCR (local)

# Tesseract specific import
try:
    import pytesseract
    # Cấu hình đường dẫn cài đặt Tesseract-OCR trên Windows
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

class ExcelPrintRedirector:
    def __init__(self, ws=None):
        self.ws = ws
        self.terminal = sys.stdout
        self.buffer = ""
        self.encoding = getattr(sys.stdout, 'encoding', 'utf-8')

    def write(self, message):
        self.terminal.write(message)
        self.buffer += message
        while '\n' in self.buffer:
            line, self.buffer = self.buffer.split('\n', 1)
            line = line.replace('\r', '')
            if self.ws:
                self.ws.append([line])
                cell = self.ws.cell(row=self.ws.max_row, column=1)
                if "✅" in line:
                    cell.font = openpyxl.styles.Font(color="00B050", bold=True)
                elif "❌" in line:
                    cell.font = openpyxl.styles.Font(color="FF0000", bold=True)
                elif "⚠️" in line:
                    cell.font = openpyxl.styles.Font(color="E26B0A", bold=True)
                elif "====" in line or "---" in line or line.strip().startswith("["):
                    cell.font = openpyxl.styles.Font(bold=True)

    def flush(self):
        self.terminal.flush()

    def set_sheet(self, ws):
        self.ws = ws
        if self.ws:
            self.ws.column_dimensions['A'].width = 150

def get_result_sheet_name(evidence_sheet_name):
    if "共通" in evidence_sheet_name:
        return "テスト計画書兼結果報告書(共通)"
    elif "個別" in evidence_sheet_name:
        return "テスト計画書兼結果報告書(個別)"
    return None

def clean_text(text):
    """Làm sạch khoảng trắng để so sánh header chính xác hơn"""
    if text is None: return ""
    return str(text).replace(' ', '').replace('　', '').strip()

def check_log_evidence(file_path):
    if not os.path.exists(file_path):
        print(f"Lỗi: Không tìm thấy file:\n{file_path}")
        return

    print(f"Đang đọc dữ liệu từ file: {os.path.basename(file_path)} ...\n")
    try:
        # Dùng data_only=True để lấy giá trị thực thay vì công thức
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Lỗi khi mở file Excel: {e}")
        return

    print(f"================ TRẠNG THÁI OCR ================\n{'✅ Đã bật (Sẵn sàng đọc ảnh)' if HAS_OCR else '⚠️ Đã tắt (Vui lòng cài pytesseract, Pillow và Tesseract-OCR)'}\n")

    print("================ KIỂM TRA TRẠNG THÁI MỞ FILE ================")
    active_sheet = wb.active
    first_sheet = wb.worksheets[0]
    
    if active_sheet != first_sheet:
        print(f"❌ LỖI: File đang mở ở sheet '{active_sheet.title}' (yêu cầu sheet đầu tiên '{first_sheet.title}').")
    else:
        print(f"✅ File đang mở ở sheet đầu tiên '{first_sheet.title}'.")
        
    try:
        active_cell = "Unknown"
        if active_sheet.views and active_sheet.views.sheetView:
            selections = active_sheet.views.sheetView[0].selection
            if selections and len(selections) > 0:
                active_cell = selections[0].activeCell
            else:
                active_cell = "A1" # Mặc định nếu file không lưu thông tin selection
                
        if active_cell != "A1":
            print(f"❌ LỖI: Con trỏ chuột đang nằm ở ô '{active_cell}' (yêu cầu ô 'A1').")
        else:
            print(f"✅ Con trỏ chuột đang nằm ở ô 'A1'.")
    except Exception:
        pass
    print("")

    # Pre-load result sheets into memory for quick lookup
    result_sheets = {}
    expected_runs_by_sheet = {}
    for rs_name in ["テスト計画書兼結果報告書(共通)", "テスト計画書兼結果報告書(個別)"]:
        if rs_name in wb.sheetnames:
            rs = wb[rs_name]
            result_sheets[rs_name] = rs
            
            # Kiểm tra lỗi Cột B/C có dữ liệu nhưng Cột K trống
            empty_k_rows = []
            expected_runs = {}
            last_a, last_b, last_c = "", "", ""
            
            def is_int_like(s):
                return s and re.match(r'^\d+(\.0)?$', s) is not None

            for row_idx, row in enumerate(rs.iter_rows(values_only=True), start=1):
                val_a = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                val_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                val_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                
                if val_a:
                    last_a = val_a
                    if not val_b: last_b = ""
                    if not val_c: last_c = ""
                if val_b:
                    last_b = val_b
                    if not val_c: last_c = ""
                if val_c:
                    last_c = val_c
                    
                val_d = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                val_e = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                col_k_val = row[10] if len(row) > 10 else None
                
                if last_b and "項目" not in last_b and "No" not in last_b:
                    if col_k_val is not None and str(col_k_val).strip() != "":
                        m_run = re.search(r'\d+', str(col_k_val))
                        if m_run:
                            run_num = int(m_run.group())
                            # Chỉ tạo testcase ID nếu cả 3 cột A, B, C đều có giá trị và là số
                            if is_int_like(last_a) and is_int_like(last_b) and is_int_like(last_c):
                                clean_a = last_a.split('.')[0]
                                clean_b = last_b.split('.')[0]
                                clean_c = last_c.split('.')[0]
                                tc_id = f"{clean_a}-{clean_b}-{clean_c}"
                                expected_runs.setdefault(run_num, []).append(tc_id)
                    else:
                        # Tránh báo nhầm dòng bị merge hoặc dòng trống hoàn toàn
                        # Chỉ báo nếu dòng thực sự có nội dung Test (Cột B hoặc C có chữ)
                        if val_c:
                            empty_k_rows.append(row_idx)
                            
            if empty_k_rows:
                print(f"================ KIỂM TRA: {rs_name} ================")
                print(f"⚠️ CẢNH BÁO: Dòng có nội dung Test nhưng Cột K (số lần test) bị trống tại: {', '.join(map(str, empty_k_rows))}\n")
            expected_runs_by_sheet[rs_name] = expected_runs

    sheet_coverage_info = {}

    target_sheets = ['エビデンス(共通)', 'エビデンス(個別)']
    found_any_sheet = False
    global_all_groups = []

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        found_any_sheet = True
        ws = wb[sheet_name]
        print(f"================ SHEET: {sheet_name} ================")
        
        all_groups_to_analyze = []
        result_sheet_name = get_result_sheet_name(sheet_name)
        rs = result_sheets.get(result_sheet_name)

        images_by_row = {}
        for img in getattr(ws, '_images', []):
            try:
                if hasattr(img.anchor, '_from'):
                    r = img.anchor._from.row + 1
                    
                    # Trích xuất dữ liệu byte ngay lập tức để tránh lỗi I/O file bị đóng về sau
                    img_data = None
                    try:
                        img_data = img._data() if callable(getattr(img, '_data', None)) else getattr(img, '_data', None)
                    except Exception:
                        try:
                            if hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                                img.ref.seek(0)
                                img_data = img.ref.read()
                        except Exception:
                            pass
                            
                    if img_data:
                        images_by_row.setdefault(r, []).append(img_data)
            except Exception:
                pass
                
        print(f"  -> Tìm thấy {sum(len(v) for v in images_by_row.values())} ảnh trong sheet này.")

        current_group = "Unknown_Group"
        group_start_row = 1
        in_table = False
        col_idx_gyo = -1
        col_idx_nichiji = -1
        col_idx_level = -1
        col_idx_msg = -1
        
        group_data = []
        found_runs_set = set()
        group_count = 0

        def flush_group(end_row=None):
            if group_data:
                group_images = []
                limit = end_row if end_row else (ws.max_row + 1)
                for r in range(group_start_row, limit):
                    if r in images_by_row:
                        group_images.extend(images_by_row[r])
                        del images_by_row[r]
            
                all_groups_to_analyze.append({
                    'group_name': current_group,
                    'data': list(group_data),
                    'rs': rs,
                    'rs_name': result_sheet_name,
                    'images': group_images,
                    'file_name': os.path.basename(file_path),
                    'found_runs_set': found_runs_set,
                    'expected_runs': expected_runs_by_sheet.get(result_sheet_name, {})
                })
                group_data.clear()
            else:
                # Nếu group không có data log, kiểm tra xem có phải là nhóm Flow Config không
                # Được xác định nếu nằm ở sheet (共通) và tên nhóm chứa chữ 'フロー' hoặc là nhóm đầu tiên
                if sheet_name == 'エビデンス(共通)' and current_group != "Unknown_Group" and ('フロー' in current_group or group_count == 1):
                    group_images = []
                    limit = end_row if end_row else (ws.max_row + 1)
                    for r in range(group_start_row, limit):
                        if r in images_by_row:
                            group_images.extend(images_by_row[r])
                            del images_by_row[r]
                    if group_images:
                        all_groups_to_analyze.append({
                            'group_name': 'Flow_Config',
                            'data': [],
                            'rs': rs,
                            'rs_name': result_sheet_name,
                            'images': group_images,
                            'file_name': os.path.basename(file_path),
                            'found_runs_set': found_runs_set,
                            'expected_runs': expected_runs_by_sheet.get(result_sheet_name, {})
                        })

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            col_a_val = row[0]
            
            # Nhận diện nhóm chạy mới (Có giá trị ở cột A)
            if col_a_val is not None and str(col_a_val).strip() != "":
                if current_group != "Unknown_Group":
                    flush_group(end_row=row_idx)
                    in_table = False
                    
                group_count += 1
                current_group = str(col_a_val).strip().replace('\n', '').replace('\r', '')
                group_start_row = row_idx
                in_table = False

            # Đang tìm kiếm Header của bảng Log
            if not in_table:
                row_strs = [clean_text(cell) for cell in row]
                if "行" in row_strs and "日時" in row_strs and "レベル" in row_strs:
                    col_idx_gyo = row_strs.index("行")
                    col_idx_nichiji = row_strs.index("日時")
                    col_idx_level = row_strs.index("レベル")
                    col_idx_msg = row_strs.index("メッセージ") if "メッセージ" in row_strs else (col_idx_level + 1)
                    in_table = True
                    continue
                else:
                    # Thử nhận diện dòng log trực tiếp (trường hợp copy log không có header)
                    # Cấu trúc: [Số 行] -> [Thời gian 日時] -> [INFO/WARN/ERROR レベル] -> [Message]
                    found_gyo, found_nichiji, found_level = -1, -1, -1
                    for c_idx, cell_val in enumerate(row):
                        if cell_val is None: continue
                        val_str = str(cell_val).strip()
                        if not val_str: continue

                        if found_gyo == -1:
                            if re.match(r'^\d+(\.0)?$', val_str):
                                found_gyo = c_idx
                        elif found_nichiji == -1:
                            # Hỗ trợ cả định dạng chuỗi "2026-05-18T11:07:13" lẫn kiểu datetime của Excel
                            if isinstance(cell_val, (datetime.datetime, datetime.time)) or re.search(r'\d{2}:\d{2}:\d{2}', val_str):
                                found_nichiji = c_idx
                        elif found_level == -1:
                            if val_str in ["INFO", "WARN", "ERROR", "DEBUG", "TRACE", "FATAL"]:
                                found_level = c_idx
                                break
                                
                    if found_gyo != -1 and found_nichiji != -1 and found_level != -1:
                        col_idx_gyo = found_gyo
                        col_idx_nichiji = found_nichiji
                        col_idx_level = found_level
                        col_idx_msg = found_level + 1
                            
                        in_table = True

            # Đang đọc dữ liệu trong bảng Log
            if in_table:
                # Nếu độ dài của dòng không đủ để lấy cột đã lưu, nghĩa là kết thúc bảng
                if len(row) <= max(col_idx_gyo, col_idx_nichiji):
                    in_table = False
                    continue

                gyo_val = row[col_idx_gyo]
                nichiji_val = row[col_idx_nichiji]

                # Điều kiện kết thúc bảng: Cột "行" bị trống hoặc không phải là số
                if gyo_val is None or str(gyo_val).strip() == "":
                    in_table = False
                    continue

                try:
                    gyo_num = int(float(str(gyo_val).strip()))
                    
                    # Gom toàn bộ text từ cột message trở về sau 
                    # (Giải quyết triệt để vấn đề khoảng trắng tab / cột rỗng chèn giữa Level và Message)
                    msg_parts = []
                    start_col = col_idx_msg if col_idx_msg != -1 else (col_idx_level + 1 if col_idx_level != -1 else 3)
                    for c_idx in range(start_col, len(row)):
                        if row[c_idx] is not None and str(row[c_idx]).strip() != "":
                            msg_parts.append(str(row[c_idx]).strip())
                    
                    msg_val = " ".join(msg_parts)
                    group_data.append({
                        'row_idx': row_idx,
                        'gyo': gyo_num,
                        'nichiji': nichiji_val,
                        'message': msg_val
                    })
                except ValueError:
                    # Không convert được sang số -> Hết bảng
                    in_table = False

        # Quét xong sheet, nếu còn data chưa phân tích thì phân tích nốt
        flush_group(end_row=ws.max_row + 1)
            
        sheet_coverage_info[sheet_name] = {
            'expected_runs': expected_runs_by_sheet.get(result_sheet_name, {}),
            'found_runs_set': found_runs_set
        }

        for g in all_groups_to_analyze:
            g['sheet_name'] = sheet_name
            short_sheet = "共通" if "共通" in sheet_name else "個別"
            g['unique_name'] = f"{short_sheet}_{g['group_name']}"
            global_all_groups.append(g)

    gemini_result = {}
    if MANUAL_GEMINI_WEB and global_all_groups:
        export_dir = os.path.join(os.path.dirname(file_path), f"Gemini_Upload_{os.path.splitext(os.path.basename(file_path))[0]}")
        os.makedirs(export_dir, exist_ok=True)
        
        json_cache_path = os.path.join(export_dir, "gemini_result.json")
        if os.path.exists(json_cache_path):
            print(f"\n[GEMINI MANUAL MODE] ♻️ Đã tìm thấy file JSON kết quả cũ tại:\n{json_cache_path}")
            print("=> Sử dụng luôn kết quả này, bỏ qua bước xuất ảnh PDF và hỏi lại Gemini.")
            try:
                with open(json_cache_path, "r", encoding="utf-8") as f:
                    gemini_result = json.load(f)
            except Exception as e:
                print(f"❌ LỖI đọc file JSON cũ: {e}. Sẽ tiến hành xuất ảnh và hỏi lại.")
                gemini_result = {}

        group_image_mapping = {}
        all_pages_for_pdf = []
        global_page_idx = 1
        group_page_ranges = {}
        
        if not gemini_result:
            for g in global_all_groups:
                u_name = g['unique_name']
                imgs = g['images']
                if not imgs: continue
                
                safe_u_name = re.sub(r'[\\/*?:"<>|]', "_", u_name)
                
                start_page = global_page_idx
                if HAS_IMAGE_LIBS and len(imgs) > 0:
                    for img_data in imgs:
                        if img_data:
                            try:
                                im = Image.open(io.BytesIO(img_data)).convert('RGB')
                                
                                num_img = Image.new('RGB', (20, 20), color=(255, 255, 0))
                                draw = ImageDraw.Draw(num_img)
                                draw.text((6, 4), str(global_page_idx), fill=(255, 0, 0))
                                
                                nearest_filter = Image.Resampling.NEAREST if hasattr(Image, 'Resampling') else Image.NEAREST
                                num_img = num_img.resize((80, 80), nearest_filter)
                                draw_big = ImageDraw.Draw(num_img)
                                draw_big.rectangle([(0, 0), (79, 79)], outline=(255, 0, 0), width=4)
                                im.paste(num_img, (im.width - 80, 0))
                                
                                all_pages_for_pdf.append(im)
                                global_page_idx += 1
                            except Exception:
                                pass
                    group_image_mapping[u_name] = safe_u_name
                else:
                    image_data = imgs[0]
                    if image_data:
                        img_path = os.path.join(export_dir, f"{safe_u_name}.png")
                        with open(img_path, "wb") as f:
                            f.write(image_data)
                        group_image_mapping[u_name] = safe_u_name
                
                end_page = global_page_idx - 1
                if end_page >= start_page:
                    group_page_ranges[u_name] = (start_page, end_page)
            
            if all_pages_for_pdf:
                combined_pdf_path = os.path.join(export_dir, f"All_Evidence.pdf")
                all_pages_for_pdf[0].save(combined_pdf_path, "PDF", resolution=100.0, save_all=True, append_images=all_pages_for_pdf[1:])
    
            if group_image_mapping:
                print(f"\n[GEMINI MANUAL MODE] Đã xuất 1 file PDF gộp duy nhất cho toàn bộ file vào thư mục:\n{export_dir}")
                
                mapping_text = "\n".join([f"- Nhóm '{k}': Từ trang {v[0]} đến trang {v[1]}" for k, v in group_page_ranges.items()])
                
                prompt = (
                    "Tôi có 1 file PDF chứa nhiều trang ảnh chụp màn hình. Trên góc phải trên cùng của mỗi trang có ĐÁNH SỐ THỨ TỰ (1, 2, 3...) trong một ô màu vàng.\n"
                    "Dưới đây là danh sách phân bổ các trang cho từng nhóm Test:\n"
                    f"{mapping_text}\n\n"
                    "Hãy đọc các thông tin sau từ file PDF dựa trên KHOẢNG TRANG TƯƠNG ỨNG CỦA TỪNG NHÓM và trả về ĐÚNG 1 ĐOẠN JSON duy nhất (không có markdown ```json):\n"
                    "* LƯU Ý TỐI QUAN TRỌNG: BẠN PHẢI COPY CHÍNH XÁC TÊN NHÓM TRONG DANH SÁCH TRÊN ĐỂ LÀM KEY TRONG JSON. KHÔNG ĐƯỢC TỰ Ý ĐỔI TÊN HAY SỬA SỐ (VD: Tuyệt đối không tự đổi 1回目 thành 8回目).\n"
                    "1. Đối với nhóm có chữ 'Flow_Config': Tìm giá trị của biến 'P_JobnetID' (hoặc 'I_JobnetID') và 'P_InterfaceID' (hoặc 'I_InterfaceID') từ bảng 'フロー変数' (cột '初期値') HOẶC bảng '引数' (cột '値'). Điền vào 'jobnet_id' và 'interface_id'.\n"
                    "2. Đối với các nhóm test còn lại (ví dụ '共通_■1回目', '個別_■1回目'...):\n"
                    "   - BƯỚC 1: Tìm '実行開始' (start_time) và '正常終了' (duration_ms) từ bảng Status Panel (Hộp thoại フローの実行).\n"
                    "     * CẢNH BÁO QUAN TRỌNG 1: TUYỆT ĐỐI KHÔNG lấy giờ từ đồng hồ hệ thống (Taskbar ở góc phải dưới cùng màn hình). Chỉ tìm chuỗi có dạng '実行開始: HH:MM:SS'. Nhớ là giá trị này phải lấy giống ảnh. vì cần check. nếu sai thì kết quả cũng sai \n "
                    "   - BƯỚC 2: Tìm 'jobnet_id' và 'interface_id' từ MỘT TRONG HAI bảng sau:\n"
                    "     + Bảng 1: Bảng 'フロー変数' (Flow Variables). Tìm cột '初期値' (Initial value) tương ứng với biến 'P_JobnetID' (hoặc 'I_JobnetID') và 'P_InterfaceID' (hoặc 'I_InterfaceID').\n"
                    "     + Bảng 2: Bảng '引数' (Arguments) trong hộp thoại 'フローの実行'. Tìm cột '値' (Value) tương ứng với biến 'P_JobnetID' (hoặc 'I_JobnetID') và 'P_InterfaceID' (hoặc 'I_InterfaceID').\n"
                    "     * CẢNH BÁO QUAN TRỌNG 2 (CHỐNG ĐỌC NHẦM LOG): TUYỆT ĐỐI KHÔNG lấy thông tin từ các dòng log dạng text có chứa chữ 'ジョブネットID: ...' hay 'インターフェースID: ...'. CHỈ ĐƯỢC PHÉP lấy từ 2 bảng đã nêu. Nếu trong các trang của nhóm test KHÔNG có bảng nào chứa thông tin này, BẮT BUỘC để rỗng \"\".\n"
                    "   - BƯỚC 3: TÌM FILE OUTPUT: Tìm ảnh chụp File Explorer. Chú ý SỐ THỨ TỰ ở góc ảnh, ảnh File Explorer hợp lệ phải có số thứ tự LỚN HƠN (nằm sau) ảnh Status Panel. Lấy tên file và 'Date modified' (đổi sang định dạng YYYY/MM/DD HH:MM 24h) điền vào 'file_name' và 'file_modified_time'.\n"
                    "     * CHÚ Ý 1: Đọc đường dẫn (path) thư mục trên thanh địa chỉ. NẾU đường dẫn chứa chữ 'receive', ĐÓ CHẮC CHẮN LÀ FILE INPUT -> TUYỆT ĐỐI BỎ QUA và để rỗng 'file_name'.\n"
                    "     * CHÚ Ý 2 (QUAN TRỌNG): Chỉ ghi nhận file nếu 'Date modified' của file đó BẰNG HOẶC SAU thời gian '実行開始' (start_time) của ca test hiện tại. \n"
                    "     * CHÚ Ý 3: Nếu đường dẫn KHÔNG chứa 'receive' và thời gian file hợp lệ, hãy CHỌN file có 'Date modified' MỚI NHẤT (thời gian muộn hơn).\n"
                    "     * CHÚ Ý 4: Nếu 'Date modified' của file NHỎ HƠN (diễn ra trước) thời gian '実行開始' (start_time), hoặc ca test kết thúc lỗi (戻り値 khác 0) và không sinh ra file mới, hãy hiểu đó là file cũ còn sót lại -> TUYỆT ĐỐI BỎ QUA, để rỗng \"\" cho 'file_name' và 'file_modified_time'.\n"
                    "     * CHÚ Ý 5 (TỐI QUAN TRỌNG): Đọc CHÍNH XÁC nội dung hiển thị trên ảnh (nhớ chuyển đổi PM sang 24h). KHÔNG ĐƯỢC copy 'start_time' điền vào 'file_modified_time'. KHÔNG ĐƯỢC tự ý đoán mò, bịa đặt hoặc tự đồng bộ dữ liệu. Nếu không nhìn thấy rõ chữ trên ảnh, BẮT BUỘC ĐỂ RỖNG (\"\").\n"
                    "     * CHÚ Ý 6 (CHỐNG ẢO GIÁC): TUYỆT ĐỐI KHÔNG copy 'start_time', 'duration_ms', 'jobnet_id', 'interface_id', 'file_name', 'file_modified_time' hay BẤT KỲ thông tin nào từ nhóm test trước điền vào nhóm test sau. Nếu trong các trang của nhóm test hiện tại KHÔNG CÓ thông tin, BẮT BUỘC để rỗng (\"\").\n"
                    "   - BƯỚC 4: TÌM ẢNH LOG VIEWER (Nhận biết: Ảnh có chữ 'ログ設定名' hoặc 'アプリケーション' hoặc giao diện xem log). Tìm NGÀY và GIỜ của dòng cuối cùng (thường chứa chữ 'サブフローの実行が終了しました') và điền vào 'log_image_time' theo định dạng 'YYYY/MM/DD HH:MM:SS'. Nếu chỉ hiển thị giờ thì điền 'HH:MM:SS'. Nếu không có ảnh log, để rỗng \"\".\n"
                    "Nếu không tìm thấy thông tin nào, hãy để rỗng (\"\").\n\n"
                    "Cấu trúc JSON mong muốn:\n"
                    "{\n"
                    "  \"共通_Flow_Config\": {\n"
                    "    \"jobnet_id\": \"...\",\n"
                    "    \"interface_id\": \"...\"\n"
                    "  },\n"
                    "  \"<COPY Y HỆT TÊN NHÓM TRONG DANH SÁCH LÊN ĐÂY>\": {\n"
                    "    \"start_time\": \"HH:MM:SS\",\n"
                    "    \"duration_ms\": \"123\",\n"
                    "    \"jobnet_id\": \"...\",\n"
                    "    \"interface_id\": \"...\",\n"
                    "    \"file_name\": \"...\",\n"
                    "    \"file_modified_time\": \"YYYY/MM/DD HH:MM\",\n"
                    "    \"log_image_time\": \"YYYY/MM/DD HH:MM:SS\"\n"
                    "  }\n"
                    "}"
                )
                
                prompt_path = os.path.join(export_dir, "prompt.txt")
                with open(prompt_path, "w", encoding="utf-8") as f:
                    f.write(prompt)
                    
                print(f"📄 Đã tạo file chứa Prompt cho bạn tại:\n{prompt_path}\n")
                print("Vui lòng upload TẤT CẢ các ảnh này lên Gemini Web và sử dụng Prompt sau:\n")
                print("-" * 50)
                print(prompt)
                print("-" * 50)
                print("Dán kết quả JSON từ Gemini vào đây (nhấn Enter 2 lần liên tiếp để kết thúc):")
                
                # Tạm thời khôi phục lại stdout thực của terminal để giao diện console không bị ẩn do Redirector
                original_stdout = sys.stdout
                if hasattr(sys.stdout, 'terminal'):
                    sys.stdout = sys.stdout.terminal
                    
                try:
                    json_lines = []
                    empty_streak = 0
                    while True:
                        try:
                            line = input()
                            if line.strip() == '':
                                empty_streak += 1
                                if empty_streak >= 2:
                                    break
                            else:
                                empty_streak = 0
                            json_lines.append(line)
                        except EOFError:
                            break
                finally:
                    sys.stdout = original_stdout
                        
                json_str = "\n".join(json_lines).strip()
                if json_str.startswith("```json"): json_str = json_str[7:]
                if json_str.startswith("```"): json_str = json_str[3:]
                if json_str.endswith("```"): json_str = json_str[:-3]
                
                try:
                    if json_str:
                        gemini_result = json.loads(json_str)
                        print("✅ Đã parse JSON thành công!")
                        with open(json_cache_path, "w", encoding="utf-8") as f:
                            json.dump(gemini_result, f, ensure_ascii=False, indent=4)
                        print(f"💾 Đã lưu kết quả JSON vào: {json_cache_path}")
                except Exception as e:
                    print(f"❌ LỖI parse JSON: {e}")
                    print("Sẽ fallback về Tesseract OCR...")

    last_sheet = None
    for g in global_all_groups:
        sheet_name = g['sheet_name']
        if sheet_name != last_sheet:
            print(f"\n================ ĐANG PHÂN TÍCH KẾT QUẢ SHEET: {sheet_name} ================")
            last_sheet = sheet_name
            
        if MANUAL_GEMINI_WEB:
            safe_u_name = re.sub(r'[\\/*?:"<>|]', "_", g['unique_name'])
            pre_ocr = None
            if gemini_result:
                if safe_u_name in gemini_result:
                    pre_ocr = gemini_result[safe_u_name]
                else:
                    # Fallback tìm kiếm linh hoạt (Fuzzy match) vì AI thường đọc nhầm chữ hoặc tự đổi số thứ tự
                    sheet_prefix = "共通" if "共通" in g['unique_name'] else "個別"
                    
                    def get_desc(t):
                        # Bỏ tiền tố sheet
                        t = re.sub(r'^(共通_|個別_)', '', t)
                        # Xóa cụm số thứ tự và loại test ở đầu (VD: ■1回目 正常系・ -> xóa)
                        desc = re.sub(r'^■?\d+回[目系]?\s*(正常系|異常系)?\s*[・.]?\s*', '', t)
                        if not desc: 
                            desc = t # Nếu xóa xong mất hết chữ (VD test chỉ tên là "■1回目") thì lấy lại nguyên bản
                        # Đồng bộ khoảng trắng, dấu ngoặc, mũi tên
                        desc = desc.replace(' ', '').replace('　', '').replace('\n', '').replace('\r', '')
                        desc = desc.replace('(', '（').replace(')', '）').replace('->', '→').replace('−＞', '→')
                        desc = desc.replace('细', '細') # Fix AI nhầm chữ Hán
                        for hw, fw in zip("0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ", 
                                          "０１２３４５６７８９ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"):
                            desc = desc.replace(fw, hw)
                        return desc.lower()
                        
                    g_desc = get_desc(g['group_name'])
                    
                    # Ưu tiên 1: Khớp chính xác description
                    for k, v in gemini_result.items():
                        if sheet_prefix in k and get_desc(k) == g_desc:
                            pre_ocr = v
                            break
                            
                    # Ưu tiên 2: Khớp chính xác sau khi bỏ ký tự đặc biệt (Siêu linh hoạt)
                    if not pre_ocr:
                        g_pure = re.sub(r'\W+', '', g_desc)
                        for k, v in gemini_result.items():
                            if sheet_prefix in k:
                                k_pure = re.sub(r'\W+', '', get_desc(k))
                                if g_pure and k_pure and g_pure == k_pure:
                                    pre_ocr = v
                                    break

                    # Ưu tiên 3: Khớp một phần (nếu AI cắt bớt chữ dài)
                    if not pre_ocr:
                        for k, v in gemini_result.items():
                            if sheet_prefix in k:
                                k_desc = get_desc(k)
                                if (g_desc in k_desc) or (k_desc in g_desc):
                                    pre_ocr = v
                                    break

                    # Ưu tiên 4: Khớp một phần sau khi bỏ ký tự đặc biệt
                    if not pre_ocr:
                        g_pure = re.sub(r'\W+', '', g_desc)
                        for k, v in gemini_result.items():
                            if sheet_prefix in k:
                                k_pure = re.sub(r'\W+', '', get_desc(k))
                                if g_pure and k_pure and (g_pure in k_pure or k_pure in g_pure):
                                    pre_ocr = v
                                    break
            analyze_group(
                g['group_name'], g['data'], g['rs'], g['rs_name'], 
                g['images'], g['file_name'], g['found_runs_set'], pre_ocr_result=pre_ocr,
                expected_runs=g.get('expected_runs', {})
            )
        else:
            analyze_group(
                g['group_name'], g['data'], g['rs'], g['rs_name'], 
                g['images'], g['file_name'], g['found_runs_set'],
                expected_runs=g.get('expected_runs', {})
            )

    for sheet_name, info in sheet_coverage_info.items():
        print(f"\n================ TỔNG KẾT COVERAGE CHO SHEET {sheet_name} ================")
        expected_runs = info['expected_runs']
        found_runs_set = info['found_runs_set']
        
        if not expected_runs:
            print("  - ⚠️ Không có thông tin nhóm test (số lần chạy) nào được khai báo ở cột K trong Kế hoạch.")
        else:
            missing_runs = []
            for run_num in sorted(expected_runs.keys()):
                tcs = list(dict.fromkeys(expected_runs[run_num]))
                if run_num in found_runs_set:
                    print(f"  - ✅ Nhóm {run_num} (gồm TC: {', '.join(tcs)}): Đã có Log Evidence.")
                else:
                    print(f"  - ❌ LỖI: Nhóm {run_num} (gồm TC: {', '.join(tcs)}): KHÔNG TÌM THẤY Log Evidence!")
                    missing_runs.append(run_num)
            
            if not missing_runs:
                print("  => TOÀN BỘ CÁC NHÓM TEST TRONG KẾ HOẠCH ĐỀU ĐÃ CÓ EVIDENCE!")
        print("")

    if MANUAL_GEMINI_WEB and global_all_groups:
        input(f"\n✅ Đã check xong file '{os.path.basename(file_path)}'. Nhấn Enter để tiếp tục sang file tiếp theo...")

def analyze_group(group_name, data, rs, rs_name, group_images=None, file_name="", found_runs_set=None, pre_ocr_result=None, expected_runs=None):
    if expected_runs is None:
        expected_runs = {}
    if group_images is None:
        group_images = []

    if group_name == 'Flow_Config':
        print(f"\n[Kiểm tra Cấu hình Flow - Trước khi Test]:")
        if pre_ocr_result:
            jobnet_id = pre_ocr_result.get("jobnet_id")
            interface_id = pre_ocr_result.get("interface_id")

            # Trích xuất phần số từ tên file để so sánh (VD: SHRF0116 -> 0116)
            m_if = re.search(r'_([A-Za-z0-9]+)\.xlsx$', file_name)
            if_code = m_if.group(1) if m_if else ""
            m_digits = re.search(r'(\d+)$', if_code)
            digits = m_digits.group(1) if m_digits else ""

            if jobnet_id:
                print(f"  - [Ảnh Flow]: ✅ Tìm thấy P_JobnetID: '{jobnet_id}'")
                if digits and digits in jobnet_id:
                    print(f"      ✅ OK: P_JobnetID '{jobnet_id}' có chứa mã '{digits}' của file.")
                else:
                    print(f"      ❌ LỖI: P_JobnetID '{jobnet_id}' KHÔNG chứa mã '{digits}' của file.")
            else:
                print(f"  - [Ảnh Flow]: ❌ LỖI: Không tìm thấy P_JobnetID trong ảnh flow.")

            if interface_id:
                print(f"  - [Ảnh Flow]: ✅ Tìm thấy P_InterfaceID: '{interface_id}'")
                if digits and digits in interface_id:
                    print(f"      ✅ OK: P_InterfaceID '{interface_id}' có chứa mã '{digits}' của file.")
                else:
                    print(f"      ❌ LỖI: P_InterfaceID '{interface_id}' KHÔNG chứa mã '{digits}' của file.")
            else:
                print(f"  - [Ảnh Flow]: ❌ LỖI: Không tìm thấy P_InterfaceID trong ảnh flow.")
        else:
            print("  - [Ảnh Flow]: ⚠️ Không có kết quả JSON từ Gemini để đối chiếu.")
        return

    if not data:
        return

    print(f"\n[Nhóm Test]: {group_name}")
    print(f"  - Tổng số dòng log: {len(data)}")

    # 1. Check xem cột "行" có tăng dần không
    is_sorted = True
    unsorted_details = []
    for i in range(len(data) - 1):
        if data[i]['gyo'] >= data[i+1]['gyo']:
            is_sorted = False
            unsorted_details.append(f"Dòng Excel {data[i]['row_idx']} (行: {data[i]['gyo']}) -> Dòng Excel {data[i+1]['row_idx']} (行: {data[i+1]['gyo']})")

    if is_sorted:
        print("  - Trạng thái Sort (行): ✅ Đã sắp xếp tăng dần hợp lệ.")
    else:
        print("  - Trạng thái Sort (行): ❌ LỖI - Chưa tăng dần!")
        for detail in unsorted_details:
            print(f"      + Lệch tại: {detail}")

    # 2. Xử lý thời gian test
    raw_dates = set()
    for item in data:
        val = item['nichiji']
        if val is None: continue
        
        if isinstance(val, datetime.datetime):
            date_str = val.strftime("%Y/%m/%d")
        else:
            # Nếu openpyxl đọc ra dưới dạng chuỗi (String)
            val_str = str(val).strip()
            # Thay thế chữ 'T' bằng khoảng trắng (nếu có) rồi mới cắt chuỗi để loại bỏ hoàn toàn phần thời gian
            date_part = val_str.replace("T", " ").split(" ")[0]
            date_str = date_part.replace("-", "/")
        raw_dates.add(date_str)

    sorted_dates = sorted(list(raw_dates))
    if len(sorted_dates) == 1:
        print(f"  - Ngày chạy log: {sorted_dates[0]}")
    else:
        print(f"  - Ngày chạy log: Từ {sorted_dates[0]} đến {sorted_dates[-1]}")

    # 3. Check xem có sự khác nhau ở Date không (Nhiều lần test khác ngày)
    if len(raw_dates) > 1:
        print(f"  - CẢNH BÁO DATE: ⚠️ Cụm này chứa log của nhiều ngày khác nhau ({', '.join(sorted_dates)}). Có vẻ cụm này đã được chạy check nhiều lần!")

    panel_start_time = None
    panel_duration_ms = None
    jobnet_id = None
    interface_id = None
    log_image_time = None

    # Kiểm tra thời gian từ ảnh (OCR)
    if not group_images:
        print("  - [Ảnh Status Panel]: ❌ LỖI: Có log chạy nhưng không tìm thấy ảnh (Status Panel) nào ở trên bảng log!")
    elif pre_ocr_result is not None:
        panel_start_time = pre_ocr_result.get("start_time")
        panel_duration_ms = pre_ocr_result.get("duration_ms")
        jobnet_id = pre_ocr_result.get("jobnet_id")
        interface_id = pre_ocr_result.get("interface_id")
        log_image_time = pre_ocr_result.get("log_image_time")

        if jobnet_id is not None:
            print(f"  - [Ảnh Status Panel]: Tìm thấy Jobnet ID: '{jobnet_id}'")
        else:
            print(f"  - [Ảnh Status Panel]: ⚠️ Không tìm thấy Jobnet ID trong ảnh.")
            
        if interface_id is not None:
            print(f"  - [Ảnh Status Panel]: Tìm thấy Interface ID: '{interface_id}'")
            if interface_id != "":
                if file_name and interface_id.lower() in file_name.lower():
                    print(f"      ✅ OK: Interface ID '{interface_id}' có tồn tại trong tên file Excel.")
                else:
                    print(f"      ❌ LỖI: Interface ID '{interface_id}' KHÔNG tồn tại trong tên file Excel ({file_name}).")
            else:
                print(f"      ⚠️ Interface ID là chuỗi rỗng nên bỏ qua kiểm tra với tên file.")
        else:
            print(f"  - [Ảnh Status Panel]: ⚠️ Không tìm thấy Interface ID trong ảnh.")

    elif HAS_OCR and HAS_IMAGE_LIBS:
        if MANUAL_GEMINI_WEB:
            print("  - [Ảnh Status Panel]: ⚠️ Không có kết quả JSON từ Gemini. Đang dùng Tesseract OCR làm fallback (độ chính xác thấp hơn)...")
        else:
            print("  - [Ảnh Status Panel]: ⚠️ Đang dùng Tesseract OCR (độ chính xác thấp hơn)...")
        panel_start_time = None
        panel_duration_ms = None
        log_image_time = None
        for idx, image_data in enumerate(group_images):
            try:
                if image_data:
                    pil_img = Image.open(io.BytesIO(image_data))
                    
                    # Tiền xử lý ảnh: Chuyển sang ảnh xám
                    pil_img = pil_img.convert('L')
                    width, height = pil_img.size
                    
                    # CHIẾN THUẬT MỚI: Cắt riêng phần thanh Status (thường nằm sát đáy, cao khoảng 80px)
                    bottom_crop = pil_img.crop((0, max(0, height - 80), width, height))
                    
                    # Phóng to ảnh bằng 2 phương pháp (LANCZOS làm mượt, NEAREST giữ nét vuông vức chống nhìn nhầm 5 thành 6)
                    resample_filter = Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS
                    resample_nearest = Image.Resampling.NEAREST if hasattr(Image, 'Resampling') else Image.NEAREST
                    
                    full_img = pil_img.resize((width * 2, height * 2), resample_filter)
                    bottom_img_lanczos = bottom_crop.resize((width * 2, bottom_crop.size[1] * 2), resample_filter)
                    bottom_img_nearest = bottom_crop.resize((width * 2, bottom_crop.size[1] * 2), resample_nearest)
                    
                    time_regex = r'([01]?\d|2[0-3])\s*[:：]\s*([0-5]\d)\s*[:：]\s*([0-5]\d)(?:.*?(\d+)\s*ms)?'
                    
                    m = None
                    
                    # 1. ĐỌC THỜI GIAN TỪ THANH STATUS
                    # Bỏ hoàn toàn tiếng Nhật, chỉ dùng tiếng Anh (eng) để đọc số chính xác nhất
                    ocr_strategies = [
                        (bottom_img_nearest, 'eng', '--psm 7'),     # Ưu tiên 1: Cắt đáy, ảnh nét vuông (tránh nhầm 5 thành 6)
                        (bottom_img_lanczos, 'eng', '--psm 7'),     # Ưu tiên 2: Cắt đáy, ảnh mượt
                        (bottom_img_nearest, 'eng', '--psm 6'),
                        (bottom_img_lanczos, 'eng', '--psm 6'),
                        (full_img, 'eng', '--psm 11'),      # Dự phòng: Đọc rải rác toàn ảnh
                        (full_img, 'eng', '--psm 3'),       # Dự phòng: Đọc layout toàn ảnh
                    ]
                    
                    for img_variant, lang, psm in ocr_strategies:
                        text_variant = pytesseract.image_to_string(img_variant, lang=lang, config=psm)
                        m_variant = re.search(time_regex, text_variant, re.IGNORECASE | re.DOTALL)
                        if m_variant:
                            m = m_variant
                            break
                            
                    # 2. ĐỌC THÔNG SỐ JOBNET / INTERFACE ID TỪ BẢNG
                    # Dùng tiếng Anh và PSM 3 để đọc Layout, bỏ qua tiếng Nhật
                    full_text = pytesseract.image_to_string(full_img, lang='eng', config='--psm 3')

                    if m:
                        # Định dạng lại thành chuẩn HH:MM:SS
                        panel_start_time = f"{int(m.group(1)):02d}:{m.group(2)}:{m.group(3)}"
                        panel_duration_ms = m.group(4)
                        
                        # Trích xuất Jobnet ID và Interface ID thông minh bằng mảng từ vựng (Words)
                        # Giúp vượt qua các lỗi khoảng trắng / ký tự rác của OCR
                        words = re.findall(r'[A-Za-z0-9@_\-]+', full_text)
                        skip_words = {'string', 'type', 'name', 'id', 'interface', 'jobnet', 'value'}
                        
                        jobnet_id = None
                        interface_id = None
                        
                        for i, w in enumerate(words):
                            w_lower = w.lower()
                            if 'jobnet' in w_lower and not jobnet_id:
                                for next_w in words[i+1:]:
                                    nl = next_w.lower()
                                    if nl not in skip_words and len(next_w) > 3 and 'str' not in nl:
                                        jobnet_id = next_w
                                        break
                                        
                            if 'interface' in w_lower and not interface_id:
                                for next_w in words[i+1:]:
                                    nl = next_w.lower()
                                    if nl not in skip_words and len(next_w) > 3 and 'str' not in nl:
                                        interface_id = next_w
                                        break

                        if jobnet_id is not None:
                            print(f"  - [Ảnh Status Panel]: Tìm thấy Jobnet ID: '{jobnet_id}'")
                        else:
                            print(f"  - [Ảnh Status Panel]: ⚠️ Không tìm thấy Jobnet ID trong ảnh.")
                            
                        if interface_id is not None:
                            print(f"  - [Ảnh Status Panel]: Tìm thấy Interface ID: '{interface_id}'")
                            if interface_id != "":
                                if file_name and interface_id.lower() in file_name.lower():
                                    print(f"      ✅ OK: Interface ID '{interface_id}' có tồn tại trong tên file Excel.")
                                else:
                                    print(f"      ❌ LỖI: Interface ID '{interface_id}' KHÔNG tồn tại trong tên file Excel ({file_name}).")
                            else:
                                print(f"      ⚠️ Interface ID là chuỗi rỗng nên bỏ qua kiểm tra với tên file.")
                        else:
                            print(f"  - [Ảnh Status Panel]: ⚠️ Không tìm thấy Interface ID trong ảnh.")

                        break
            except Exception as e:
                print(f"      ⚠️ Lỗi khi xử lý ảnh (OCR): {e}")

    else:
        print("  - [Ảnh Status Panel]: ⚠️ Cần cài đặt thư viện OCR (pytesseract, Pillow) để đọc text từ ảnh.")
        
    if panel_start_time:
        print(f"  - [Ảnh Status Panel]: Tìm thấy giờ bắt đầu là {panel_start_time}")
        first_log_val = data[0]['nichiji']
        if first_log_val:
            first_log_time_str = ""
            if isinstance(first_log_val, datetime.datetime):
                first_log_time_str = first_log_val.strftime("%H:%M:%S")
            else:
                time_m = re.search(r'(\d{1,2}:\d{2}:\d{2})', str(first_log_val))
                if time_m:
                    first_log_time_str = time_m.group(1)
            
            if first_log_time_str:
                try:
                    fmt = "%H:%M:%S"
                    t_panel = datetime.datetime.strptime(panel_start_time, fmt).time()
                    t_log = datetime.datetime.strptime(first_log_time_str, fmt).time()
                    
                    # Tính độ trễ giây (Dung sai)
                    t_panel_dt = datetime.datetime.combine(datetime.date.today(), t_panel)
                    t_log_dt = datetime.datetime.combine(datetime.date.today(), t_log)
                    diff_seconds = (t_panel_dt - t_log_dt).total_seconds()
                    
                    if abs(diff_seconds) <= 1:
                        print(f"      ✅ OK: Thời gian log ({first_log_time_str}) khớp Thời gian ảnh ({panel_start_time}) (Lệch {int(diff_seconds)}s)")
                    else:
                        print(f"      ❌ LỖI: Thời gian log ({first_log_time_str}) lệch quá 1s so với Thời gian ảnh ({panel_start_time}) (Lệch {int(diff_seconds)}s)")
                except ValueError:
                    print(f"      ⚠️ Không thể so sánh thời gian: Log ({first_log_time_str}), Ảnh ({panel_start_time})")
    else:
        print("  - [Ảnh Status Panel]: ❌ LỖI: Không tìm thấy hoặc không đọc được thời gian '実行開始:' từ ảnh.")

    log_duration_ms = None
    log_end_time_str = None
    if data:
        last_val = data[-1]['nichiji']
        if last_val:
            if isinstance(last_val, datetime.datetime):
                log_end_time_str = last_val.strftime("%H:%M:%S")
            else:
                time_m = re.search(r'(\d{1,2}:\d{2}:\d{2})', str(last_val))
                if time_m:
                    log_end_time_str = time_m.group(1)

        # Quét ngược từ dưới lên để tìm chính xác dòng log kết thúc subflow
        for item in reversed(data):
            msg = str(item.get('message', ''))
            if "サブフローの実行が終了しました" in msg:
                m_log_dur = re.search(r'\[\s*(\d+)(?:\s*ms)?\s*\]', msg, re.IGNORECASE)
                if m_log_dur:
                    log_duration_ms = m_log_dur.group(1)
                
                val = item['nichiji']
                if val:
                    if isinstance(val, datetime.datetime):
                        log_end_time_str = val.strftime("%H:%M:%S")
                    else:
                        time_m = re.search(r'(\d{1,2}:\d{2}:\d{2})', str(val))
                        if time_m:
                            log_end_time_str = time_m.group(1)
                    break

    if panel_duration_ms and data:
        if log_duration_ms:
            try:
                if int(log_duration_ms) < int(panel_duration_ms):
                    print(f"      ✅ OK: Duration Log Main ({log_duration_ms}ms) < Duration Ảnh ({panel_duration_ms}ms)")
                else:
                    print(f"      ❌ LỖI: Duration Log Main ({log_duration_ms}ms) >= Duration Ảnh ({panel_duration_ms}ms)")
            except ValueError:
                pass
        else:
            print(f"      ⚠️ Không tìm thấy dòng log 'サブフローの実行が終了しました' chứa thời gian '[Nms]' hoặc '[N]' để so sánh Duration.")

    if log_image_time:
        print(f"  - [Ảnh Log Viewer]: Tìm thấy thời gian kết thúc trong ảnh log là '{log_image_time}'")
        
        img_date_str = None
        img_time_str = log_image_time
        if " " in log_image_time:
            parts = log_image_time.split(" ")
            img_date_str = parts[0].replace("-", "/")
            img_time_str = parts[1]
            
        if log_end_time_str:
            try:
                fmt = "%H:%M:%S"
                t_img = datetime.datetime.strptime(img_time_str, fmt).time()
                t_txt = datetime.datetime.strptime(log_end_time_str, fmt).time()
                
                if t_img == t_txt:
                    print(f"      ✅ OK: Giờ kết thúc trong ảnh Log ({img_time_str}) KHỚP với log text ({log_end_time_str}).")
                else:
                    print(f"      ❌ LỖI: Giờ kết thúc trong ảnh Log ({img_time_str}) LỆCH với log text ({log_end_time_str}).")
            except ValueError:
                print(f"      ⚠️ Không thể so sánh giờ kết thúc: Ảnh ({img_time_str}), Log ({log_end_time_str})")
                
            if img_date_str and sorted_dates:
                log_date_str = sorted_dates[-1]
                if img_date_str == log_date_str:
                    print(f"      ✅ OK: Ngày trong ảnh Log ({img_date_str}) KHỚP với ngày log text ({log_date_str}).")
                else:
                    print(f"      ❌ LỖI: Ngày trong ảnh Log ({img_date_str}) LỆCH với ngày log text ({log_date_str}).")
        else:
            print(f"      ⚠️ Không tìm thấy thời gian kết thúc trong log text để đối chiếu.")
            
        if panel_start_time:
            try:
                t_start = datetime.datetime.strptime(panel_start_time, "%H:%M:%S").time()
                t_img = datetime.datetime.strptime(img_time_str, "%H:%M:%S").time()
                if t_img >= t_start:
                    print(f"      ✅ OK: Giờ ảnh Log ({img_time_str}) >= Giờ bắt đầu ({panel_start_time}).")
                else:
                    print(f"      ❌ LỖI: Giờ ảnh Log ({img_time_str}) < Giờ bắt đầu ({panel_start_time}).")
            except ValueError:
                pass
    else:
        print(f"  - [Ảnh Log Viewer]: ⚠️ CẢNH BÁO: Không tìm thấy ảnh chụp màn hình Log (không có chữ ログ設定名 / アプリケーション) hoặc không đọc được thời gian.")

    if jobnet_id is not None and data:
        jobnet_in_log = False
        log_jobnet_found_val = None
        
        # Chỉ lấy phần trước '@' của Jobnet ID trong ảnh để so sánh
        jobnet_id_compare = jobnet_id.split('@')[0] if jobnet_id else ""
        
        for item in data:
            msg = str(item.get('message', ''))
            if "ジョブネットID：" in msg:
                m_log_jobnet = re.search(r'ジョブネットID：([^\s　]*)', msg)
                if m_log_jobnet:
                    log_jobnet_found_val = m_log_jobnet.group(1)
                    log_jobnet_compare = log_jobnet_found_val.split('@')[0] if log_jobnet_found_val else ""
                    if log_jobnet_compare == jobnet_id_compare:
                        jobnet_in_log = True
                        break
            elif jobnet_id_compare and jobnet_id_compare in msg:
                jobnet_in_log = True
                break
                
        if jobnet_in_log:
            print(f"      ✅ OK: Jobnet ID '{jobnet_id}' khớp với dữ liệu trong Log.")
        else:
            if log_jobnet_found_val is not None:
                print(f"      ❌ LỖI: Jobnet ID trong ảnh là '{jobnet_id}', nhưng trong Log là '{log_jobnet_found_val}'.")
            else:
                print(f"      ❌ LỖI: Không tìm thấy dòng log chứa 'ジョブネットID：' hoặc Jobnet ID '{jobnet_id}' trong Log.")

    file_name_ocr = pre_ocr_result.get("file_name") if pre_ocr_result else None
    file_modified_time = pre_ocr_result.get("file_modified_time") if pre_ocr_result else None

    # Kiểm tra File name trong Log text
    log_filename_found = None
    if data:
        for item in data:
            msg = str(item.get('message', ''))
            m_file = re.search(r'ファイル名[：:]([^\s　]+)', msg)
            if m_file:
                log_filename_found = m_file.group(1)
                break
                
    if log_filename_found:
        print(f"  - [Kiểm tra File]: Log text báo có tạo ra file '{log_filename_found}'.")
        if not file_name_ocr:
            print(f"      ❌ LỖI: Log báo có tạo file nhưng không tìm thấy ảnh chụp File Explorer chứa file này!")
        elif log_filename_found.lower() == file_name_ocr.lower():
            print(f"      ✅ OK: Tên file trong ảnh File Explorer ({file_name_ocr}) KHỚP với tên file trong Log text.")
        else:
            print(f"      ❌ LỖI: Tên file trong ảnh File Explorer ({file_name_ocr}) KHÔNG KHỚP với tên file trong Log text ({log_filename_found}).")

    if file_name_ocr and file_modified_time:
        print(f"  - [Ảnh File Explorer]: Tìm thấy file '{file_name_ocr}' có Date modified là '{file_modified_time}'")
        if log_end_time_str and sorted_dates:
            try:
                log_date_str = sorted_dates[-1]
                log_datetime_str = f"{log_date_str} {log_end_time_str[:5]}"
                
                file_modified_hm = file_modified_time[:16]
                if log_datetime_str == file_modified_hm:
                    print(f"      ✅ OK: Thời gian xuất file '{file_name_ocr}' ({file_modified_time}) KHỚP với log ({log_datetime_str}).")
                else:
                    print(f"      ❌ LỖI: Thời gian xuất file '{file_name_ocr}' ({file_modified_time}) LỆCH với log ({log_datetime_str}).")
            except Exception as e:
                print(f"      ⚠️ Không thể so sánh thời gian xuất file: {e}")

    # Bóc tách số lần chạy từ tên nhóm (Ví dụ: ■1回目 -> 1)
    m_run = re.search(r'(\d+)回目', group_name)
    extracted_run = None
    base_group_name = group_name
    if m_run:
        extracted_run = int(m_run.group(1))
        base_group_name = re.sub(r'^■?\s*\d+回目\s*', '', group_name).strip()

    # 5. Check testcase coverage in log
    if extracted_run and expected_runs:
        expected_tcs_for_run = set(expected_runs.get(extracted_run, []))
        if expected_tcs_for_run:
            found_tcs_in_log = set()
            for item in data:
                msg = str(item.get('message', ''))
                # Find all occurrences of 'd-d-d' pattern
                found_tcs_in_log.update(re.findall(r'\b\d+-\d+-\d+\b', msg))

            print(f"  - [Đối chiếu Testcase trong Log]:")
            missing_tcs = expected_tcs_for_run - found_tcs_in_log
            extra_tcs = found_tcs_in_log - expected_tcs_for_run

            if not missing_tcs and not extra_tcs:
                print(f"      ✅ OK: Toàn bộ {len(expected_tcs_for_run)} testcase dự kiến ({', '.join(sorted(list(expected_tcs_for_run)))}) đều có trong log.")
            else:
                if missing_tcs:
                    None
                    # print(f"      ❌ LỖI: Thiếu {len(missing_tcs)} testcase trong log: {', '.join(sorted(list(missing_tcs)))}")
                if extra_tcs:
                    print(f"      ⚠️ CẢNH BÁO: Phát hiện testcase không có trong kế hoạch cho nhóm này: {', '.join(sorted(list(extra_tcs)))}")
        else:
            print(f"  - [Đối chiếu Testcase trong Log]: ⚠️ Không tìm thấy testcase nào được định nghĩa cho nhóm {extracted_run} trong Kế hoạch.")

    # 4. Check result sheet
    if rs:
        found_rows = []
        last_b, last_c, last_d = "", "", ""
        
        for row_idx, row in enumerate(rs.iter_rows(values_only=True), start=1):
            val_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            val_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            val_d = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            
            if val_b: last_b = val_b
            if val_c: last_c = val_c
            if val_d: last_d = val_d
            
            # Yêu cầu: Chỉ check nếu cột C (項目) của dòng đó có giá trị
            if not val_c:
                continue
                
            if group_name in [val_b, val_c, val_d]:
                found_rows.append(row_idx)
                continue
                
            if base_group_name and base_group_name in [val_b, val_c, val_d]:
                found_rows.append(row_idx)
                continue
                
            tc_id = f"{last_b}-{last_c}-{last_d}"
            if last_b and last_c and last_d and (group_name == tc_id or base_group_name == tc_id):
                found_rows.append(row_idx)
                continue
                
        # Fallback (Dự phòng): Nếu tìm theo tên không ra, tìm theo số lần chạy ở cột K (áp dụng cho cả 共通 và 個別)
        if not found_rows:
            last_c_fallback = "" # Xử lý cho trường hợp cột C bị merge
            target_run_num = extracted_run
            
            for row_idx, row in enumerate(rs.iter_rows(values_only=True), start=1):
                val_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                if val_c:
                    last_c_fallback = val_c

                col_k_val = row[10] if len(row) > 10 else None
                
                if last_c_fallback and col_k_val is not None and str(col_k_val).strip() != "":
                    # Bóc tách số từ cột K (Ví dụ: "1" -> 1, "1回目" -> 1)
                    m_run = re.search(r'\d+', str(col_k_val))
                    run_num = int(m_run.group()) if m_run else None

                    if run_num is not None:
                        if target_run_num is not None:
                            if run_num == target_run_num:
                                found_rows.append(row_idx)
                        else:
                            target_run_num = run_num
                            found_rows.append(row_idx)

        if found_rows:
            print(f"  - [Đối chiếu {rs_name}]: Tìm thấy {len(found_rows)} dòng ({', '.join(map(str, found_rows))})")
            
            invalid_k_rows = []
            invalid_g_rows = []
            empty_g_rows = []
            
            for f_row in found_rows:
                row_data = list(rs.iter_rows(min_row=f_row, max_row=f_row, values_only=True))[0]
                col_g_date = row_data[6] if len(row_data) > 6 else None
                col_i_date = row_data[8] if len(row_data) > 8 else None
                col_k_runs = row_data[10] if len(row_data) > 10 else None
                
                # Check số lần chạy (Cột K)
                expected_runs = None
                if col_k_runs is not None and str(col_k_runs).strip() != "":
                    m = re.search(r'\d+', str(col_k_runs))
                    if m:
                        expected_runs = int(m.group())
                        
                if expected_runs is not None and extracted_run is not None:
                    if extracted_run != expected_runs:
                        invalid_k_rows.append((f_row, expected_runs))
                        
                # Check Ngày (Ưu tiên cột I, nếu trống lấy cột G)
                date_source_val = col_i_date
                source_date_col_name = "I"
                if date_source_val is None or str(date_source_val).strip() == "":
                    date_source_val = col_g_date
                    source_date_col_name = "G"

                expected_date = None
                if date_source_val is not None and str(date_source_val).strip() != "":
                    if isinstance(date_source_val, datetime.datetime):
                        expected_date = date_source_val.strftime("%Y/%m/%d")
                    else:
                        date_str = str(date_source_val).strip()
                        date_part = date_str.replace("T", " ").split(" ")[0]
                        expected_date = date_part.replace("-", "/")
                        
                if expected_date:
                    if expected_date not in sorted_dates:
                        invalid_g_rows.append((f_row, expected_date, source_date_col_name))
                else:
                    empty_g_rows.append(f_row)

            # In kết quả số lần chạy (Cột K)
            if invalid_k_rows:
                for r, exp in invalid_k_rows:
                    print(f"      + Số lần test (Dòng {r}): ❌ LỆCH! Log là lần {extracted_run}, nhưng Cột K ghi {exp}")
            elif extracted_run is not None:
                print(f"      + Số lần test (Cột K): ✅ Tất cả {len(found_rows)} dòng đều khớp (Lần {extracted_run})")
                
            # In kết quả Ngày test (Cột I/G)
            if invalid_g_rows:
                for r, exp_d, col_name in invalid_g_rows:
                    print(f"      + Ngày test (Dòng {r}): ❌ LỆCH! Cột {col_name} ghi '{exp_d}' nhưng Log chạy ngày {', '.join(sorted_dates)}")
            
            if empty_g_rows:
                print(f"      + Ngày test (Cột I/G): ⚠️ Trống tại các dòng: {', '.join(map(str, empty_g_rows))}")
                
            if not invalid_g_rows and len(found_rows) > len(empty_g_rows):
                print(f"      + Ngày test (Cột I/G): ✅ Khớp ngày ({', '.join(sorted_dates)}) cho tất cả các dòng có dữ liệu")
        else:
            print(f"  - [Đối chiếu]: ⚠️ Không tìm thấy testcase '{group_name}' trong sheet {rs_name}")

    if target_run_num is not None and found_runs_set is not None:
        found_runs_set.add(target_run_num)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Tool kiểm tra Evidence Log (Sort tăng dần, check thời gian) cho file hoặc folder.")
    parser.add_argument("path", help="Đường dẫn file excel hoặc thư mục chứa các file excel.")
    parser.add_argument("--start", type=int, default=1, help="Vị trí file bắt đầu xử lý (1-based, mặc định: 1)")
    parser.add_argument("--out-excel", type=str, default="Evidence_Check_Report.xlsx", help="Tên file Excel xuất báo cáo")
    args = parser.parse_args()
    
    input_path = args.path
    out_excel_path = args.out_excel

    report_wb = openpyxl.Workbook()
    if report_wb.active:
        report_wb.remove(report_wb.active)

    redirector = ExcelPrintRedirector(None)
    sys.stdout = redirector

    try:
        if not os.path.exists(input_path):
            print(f"Lỗi: Đường dẫn không tồn tại: {input_path}")
        elif os.path.isfile(input_path):
            if args.start != 1:
                print(f"⚠️ Cảnh báo: Tham số --start={args.start} bị bỏ qua vì bạn đang truyền vào 1 file duy nhất thay vì 1 thư mục.")
            
            sheet_title = os.path.splitext(os.path.basename(input_path))[0][:31]
            ws = report_wb.create_sheet(title=sheet_title)
            redirector.set_sheet(ws)
            check_log_evidence(input_path)
        elif os.path.isdir(input_path):
            print(f"--- Đang quét thư mục '{input_path}' để tìm file .xlsx ---")
            excel_files = []
            for root, dirs, files in os.walk(input_path):
                for file in files:
                    if file.endswith('.xlsx') and not file.startswith('~$'):
                        excel_files.append(os.path.join(root, file))
            
            if not excel_files:
                print("Không tìm thấy file .xlsx nào trong thư mục được chỉ định.")
            else:
                excel_files.sort()
                total_files = len(excel_files)
                print(f"Tìm thấy {total_files} file.")
                start_idx = max(0, args.start - 1)
                
                if start_idx >= total_files:
                    print(f"Lỗi: Tham số --start ({args.start}) lớn hơn tổng số file ({total_files}).")
                else:
                    if start_idx > 0:
                        print(f"Bỏ qua {start_idx} file đầu tiên. Bắt đầu xử lý từ file thứ {args.start}...\n")
                    else:
                        print("Bắt đầu xử lý...\n")
                        
                    for i, file_path in enumerate(excel_files[start_idx:]):
                        current_idx = start_idx + i + 1
                        
                        sheet_title = os.path.splitext(os.path.basename(file_path))[0][:31]
                        base_title = sheet_title
                        counter = 1
                        while sheet_title in report_wb.sheetnames:
                            suffix = f"_{counter}"
                            sheet_title = base_title[:31-len(suffix)] + suffix
                            counter += 1
                            
                        ws = report_wb.create_sheet(title=sheet_title)
                        redirector.set_sheet(ws)
                        
                        print(f"\n{'='*25} [{current_idx}/{total_files}] ĐANG XỬ LÝ FILE: {os.path.basename(file_path)} {'='*25}")
                        check_log_evidence(file_path)
    finally:
        sys.stdout = redirector.terminal
        if len(report_wb.sheetnames) > 0:
            try:
                report_wb.save(out_excel_path)
                print(f"\n✅ Đã xuất báo cáo tổng hợp ra file Excel: {os.path.abspath(out_excel_path)}")
            except Exception as e:
                print(f"\n❌ Lỗi khi lưu file báo cáo Excel: {e}")